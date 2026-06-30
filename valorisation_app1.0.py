"""
Outil de valorisation par adresse — version Streamlit (locale)
----------------------------------------------------------------
Installation (une seule fois) :
    pip install streamlit pandas openpyxl requests

Lancement :
    streamlit run app.py

Charge le fichier Excel au format du gabarit (onglet COMPARABLES),
géocode via l'API Adresse du gouvernement (api-adresse.data.gouv.fr,
gratuite, sans clé), filtre par rayon et calcule une valeur
basse/cible/haute pondérée.
"""

import io
import math
import re
import time
from datetime import datetime

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import live_search

st.set_page_config(page_title="Valorisation par adresse", page_icon="🏢", layout="wide")

BAN_URL = "https://api-adresse.data.gouv.fr/search/"


# ---------------------------------------------------------------- géocodage
@st.cache_data(show_spinner=False)
def geocode(query: str):
    """Retourne (lat, lon, label) ou None. Mis en cache pour ne pas re-requêter."""
    if not query or not str(query).strip():
        return None
    try:
        r = requests.get(BAN_URL, params={"q": query, "limit": 1}, timeout=8)
        r.raise_for_status()
        feats = r.json().get("features", [])
        if not feats:
            return None
        f = feats[0]
        lon, lat = f["geometry"]["coordinates"]
        return lat, lon, f["properties"].get("label", query)
    except requests.RequestException:
        return None


def haversine_m(lat1, lon1, lat2, lon2):
    R = 6371000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlmb / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def to_num(v):
    try:
        if v is None or v == "":
            return None
        return float(str(v).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return None


# ------------------------------------------------- extraction type de bail / preneur
BAIL_RE = re.compile(r"bail\s+(commercial|d[ée]rogatoire|professionnel)\s*(\d/\d/\d)?", re.I)
PRENEUR_RE = re.compile(r"lou[ée]\s+(?:à|a)\s+(?:la\s+soci[ée]t[ée]\s+)?([A-ZÀ-Ü][\w\-' ]{2,40})", re.I)


def extract_type_bail(row_):
    obs = str(row_.get("Observations", "") or "")
    m = BAIL_RE.search(obs)
    if m:
        return f"Bail {m.group(1)}{' ' + m.group(2) if m.group(2) else ''}".strip()
    opv = row_.get("Operation")
    if pd.notna(opv) and opv not in ("", "Non disponible"):
        return str(opv)
    return "Non disponible"


def extract_preneur_etat(row_):
    obs = str(row_.get("Observations", "") or "")
    m = PRENEUR_RE.search(obs)
    if m:
        return m.group(1).strip()
    etat = row_.get("Etat")
    if pd.notna(etat) and etat not in ("", "Non disponible"):
        return str(etat)
    return "Non disponible"


def build_export_table(res_df: pd.DataFrame) -> pd.DataFrame:
    """Transforme le tableau de résultats au format demandé :
    Date | Adresse | Type de bail | Surfaces (m²) | Loyer € H.T-H.C/m²/an | Preneur / Etat locaux"""
    rows = []
    for _, r in res_df.iterrows():
        date_val = r.get("Date_transaction")
        if pd.isna(date_val) or date_val in ("", None):
            date_val = r.get("Date_collecte", "Non disponible")

        adresse = str(r.get("Adresse", "") or "").strip()
        commune = str(r.get("Commune", "") or "").strip()
        adresse_complete = (f"{adresse}, {commune}" if commune and commune not in adresse
                             else (adresse or "Non disponible"))

        loyer = to_num(r.get("Loyer_facial_eur_m2_an"))

        rows.append({
            "Date": date_val if date_val not in (None, "") else "Non disponible",
            "Adresse": adresse_complete,
            "Type de bail": extract_type_bail(r),
            "Surfaces (m²)": to_num(r.get("Surface_m2")),
            "Loyer € H.T-H.C/m²/an": loyer if loyer is not None else "Non disponible",
            "Preneur / Etat locaux": extract_preneur_etat(r),
        })
    return pd.DataFrame(rows)


def export_to_excel_bytes(export_df: pd.DataFrame, title: str = "Extraction comparables") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Extraction"

    HEAD = PatternFill("solid", start_color="1F3864")
    thin = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    cols = list(export_df.columns)
    widths = [14, 38, 20, 14, 22, 26]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    t = ws.cell(1, 1, title)
    t.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill = HEAD
    t.alignment = CENTER
    ws.row_dimensions[1].height = 24

    for j, (h, w) in enumerate(zip(cols, widths), start=1):
        c = ws.cell(2, j, h)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = HEAD
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[2].height = 30

    for i, row_ in export_df.iterrows():
        r = i + 3
        for j, col in enumerate(cols, start=1):
            val = row_[col]
            c = ws.cell(r, j, val)
            c.border = BORDER
            c.font = Font(name="Arial", size=9)
            c.alignment = LEFT if col == "Adresse" else CENTER
            if col == "Loyer € H.T-H.C/m²/an" and isinstance(val, (int, float)):
                c.number_format = '#,##0 €;(#,##0 €);"-"'
            if col == "Surfaces (m²)" and isinstance(val, (int, float)):
                c.number_format = "#,##0"

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------- UI : chargement
st.title("🏢 Valorisation par adresse")
st.caption("Tertiaire — bureaux · activités · commerce · mixte — géocodage réel via l'API Adresse (data.gouv.fr)")

with st.sidebar:
    st.header("1 — Source des comparables")
    mode = st.radio("Mode", ["Charger un fichier Excel", "Rechercher en direct (expérimental)"])

    file = None
    live_commune, live_cp, live_op, live_type = None, None, None, None
    if mode == "Charger un fichier Excel":
        file = st.file_uploader("Fichier Excel (gabarit, onglet COMPARABLES)", type=["xlsx"])
    else:
        st.caption("⚠️ Best-effort : peut ne rien retourner si les sites rendent leur "
                    "contenu en JavaScript. Voir la console pour le diagnostic.")
        live_commune = st.text_input("Commune", value="Nantes")
        live_cp = st.text_input("Code postal", value="44000")
        live_op = st.selectbox("Opération recherchée", ["Location", "Vente"])
        live_type = st.selectbox("Type de bien", ["Bureaux", "Activités", "Commerce"])
        run_live = st.button("Lancer la recherche en direct")

    st.header("2 — Actif à valoriser")
    address = st.text_input("Adresse", placeholder="ex. 5 rue de la Tour, 44200 Nantes")
    target = geocode(address) if address else None
    if address and not target:
        st.error("Adresse introuvable ou réseau indisponible. Vérifie l'orthographe / ajoute la commune.")
    elif target:
        st.success(f"✓ {target[2]}")

    st.header("3 — Critères")
    op = st.selectbox("Opération", ["Location", "Vente", "Tous"])
    asset_type = st.selectbox("Type d'actif", ["Tous", "Bureaux", "Activités", "Commerce", "Mixte"])
    surface_target = st.number_input("Surface de l'actif (m²) — optionnel", min_value=0, value=0, step=10)
    radius_km = st.slider("Rayon de recherche (km)", 0.2, 15.0, 1.5, 0.1)

    st.subheader("Pondération du score")
    w_prox = st.slider("Proximité", 0, 100, 45)
    w_sim = st.slider("Similarité (surface)", 0, 100, 35)
    w_rec = st.slider("Récence", 0, 100, 20)

if mode == "Charger un fichier Excel":
    if not file:
        st.info("Charge ton fichier Excel dans la barre latérale pour démarrer.")
        st.stop()
    try:
        df = pd.read_excel(file, sheet_name="COMPARABLES", skiprows=2)
    except Exception:
        df = pd.read_excel(file, skiprows=2)
    df = df[df["Commune"].notna() & df["Surface_m2"].notna()].reset_index(drop=True)
    st.write(f"**{len(df)}** référence(s) chargée(s) depuis le fichier.")
else:
    if "live_results" not in st.session_state:
        st.session_state.live_results = []
    if run_live:
        with st.spinner(f"Recherche en direct sur {live_commune}…"):
            res_bl = live_search.search_bureauxlocaux(live_commune, live_cp, live_op, live_type)
            res_geo = live_search.search_geolocaux(live_commune, live_cp, live_op, live_type)
            st.session_state.live_results = res_bl + res_geo
        if not st.session_state.live_results:
            st.error(
                "Aucune donnée extraite. Cause la plus probable : ces sites chargent leurs "
                "annonces en JavaScript après le chargement de la page, ce qui empêche "
                "l'extraction par requête HTTP simple. Solution : utiliser le mode "
                "'Charger un fichier Excel' avec des données collectées manuellement, "
                "ou demander une version Playwright (navigateur simulé) du scraper."
            )
            st.stop()
        else:
            st.success(f"{len(st.session_state.live_results)} annonce(s) extraite(s) en direct.")
    if not st.session_state.live_results:
        st.info("Renseigne les critères et lance la recherche en direct dans la barre latérale.")
        st.stop()
    df = pd.DataFrame(st.session_state.live_results)
    st.write(f"**{len(df)}** référence(s) extraite(s) en direct (non vérifiées, à valider).")
    st.dataframe(df, use_container_width=True, hide_index=True)

if not target:
    st.warning("Renseigne et valide une adresse cible pour lancer l'analyse.")
    st.stop()

t_lat, t_lon, t_label = target
radius_m = radius_km * 1000

# ---------------------------------------------------------------- géocodage des comps
progress = st.progress(0.0, text="Géocodage des comparables…")
lats, lons, dists = [], [], []
for i, row_ in df.iterrows():
    lat = to_num(row_.get("Latitude"))
    lon = to_num(row_.get("Longitude"))
    if lat is None or lon is None:
        q = f"{row_.get('Adresse', '')} {row_.get('Code_postal', '')} {row_.get('Commune', '')}"
        g = geocode(q)
        if g:
            lat, lon, _ = g
        time.sleep(0.05)  # courtoisie API publique
    lats.append(lat)
    lons.append(lon)
    dists.append(haversine_m(t_lat, t_lon, lat, lon) if (lat and lon) else None)
    progress.progress((i + 1) / max(len(df), 1), text=f"Géocodage… {i+1}/{len(df)}")
progress.empty()

df["_lat"], df["_lon"], df["_dist"] = lats, lons, dists

# ---------------------------------------------------------------- filtrage + score
def passes(row_):
    if row_["_dist"] is None or row_["_dist"] > radius_m:
        return False
    if op != "Tous" and pd.notna(row_.get("Operation")) and row_.get("Operation") != op:
        return False
    if asset_type != "Tous" and pd.notna(row_.get("Type_actif")) and row_.get("Type_actif") != asset_type:
        return False
    return True

res = df[df.apply(passes, axis=1)].copy()

if res.empty:
    st.warning("Aucun comparable dans ce rayon avec ces critères. Élargis le rayon ou change l'opération.")
    st.stop()

now = datetime.now()
def score_row(row_):
    p_prox = max(0.0, 1 - row_["_dist"] / radius_m)
    sc = to_num(row_.get("Surface_m2"))
    if surface_target and sc:
        p_sim = min(sc, surface_target) / max(sc, surface_target)
    else:
        p_sim = 0.5
    p_rec = 0.5
    dt = pd.to_datetime(row_.get("Date_transaction"), errors="coerce")
    if pd.notna(dt):
        months = (now - dt.to_pydatetime()).days / 30.4
        p_rec = max(0.0, 1 - months / 36)
    wsum = (w_prox + w_sim + w_rec) or 1
    return (w_prox * p_prox + w_sim * p_sim + w_rec * p_rec) / wsum

res["_score"] = res.apply(score_row, axis=1)
res = res.sort_values("_score", ascending=False)

field = "Prix_vente_eur_m2" if op == "Vente" else "Loyer_facial_eur_m2_an"
vals = res[field].apply(to_num).dropna()
weights = res.loc[vals.index, "_score"]

# ---------------------------------------------------------------- résultats
st.divider()
if len(vals) == 0:
    st.warning(f"Comparables trouvés mais aucun n'a de valeur renseignée dans la colonne {field}.")
    st.stop()

cible = (vals * weights).sum() / weights.sum()
basse, haute, med = vals.min(), vals.max(), vals.median()
unit = "€/m²" if op == "Vente" else "€/m²/an"

c1, c2, c3, c4 = st.columns(4)
c1.metric("Valeur basse" if op == "Vente" else "Loyer bas", f"{basse:,.0f} {unit}".replace(",", " "))
c2.metric("Valeur cible" if op == "Vente" else "Loyer cible", f"{cible:,.0f} {unit}".replace(",", " "))
c3.metric("Valeur haute" if op == "Vente" else "Loyer haut", f"{haute:,.0f} {unit}".replace(",", " "))
c4.metric("Comparables retenus", len(res))

n = len(vals)
conf = "🟢 Élevée" if n >= 8 else ("🟡 Moyenne" if n >= 4 else "🔴 Faible")
st.caption(f"Médiane : {med:,.0f} {unit} — Fiabilité : {conf} (basée sur {n} comparable(s) avec valeur renseignée)".replace(",", " "))
if n < 4:
    st.error("Peu de comparables exploitables : élargis le rayon ou enrichis la base.")

st.subheader(f"Comparables retenus ({len(res)})")
show_cols = ["_score", "_dist", "Adresse", "Commune", "Type_actif", "Surface_m2",
             field, "Date_transaction", "Fiabilite", "Source", "Lien"]
show_cols = [c for c in show_cols if c in res.columns]
display = res[show_cols].copy()
display["_score"] = (display["_score"] * 100).round(0).astype(int)
display["_dist"] = display["_dist"].round(0).astype(int)
display = display.rename(columns={"_score": "Score", "_dist": "Distance (m)"})
st.dataframe(display, use_container_width=True, hide_index=True)

st.divider()
st.subheader("Export pour rapport")
export_df = build_export_table(res)
st.dataframe(export_df, use_container_width=True, hide_index=True)

col_a, col_b = st.columns(2)
with col_a:
    st.download_button(
        "📥 Télécharger l'extraction (Excel)",
        export_to_excel_bytes(export_df, title=f"Extraction comparables — {t_label}"),
        file_name="extraction_comparables.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
with col_b:
    st.download_button(
        "Télécharger le détail complet (CSV)",
        display.to_csv(index=False).encode("utf-8"),
        file_name="comparables_detail_complet.csv",
        mime="text/csv",
    )
