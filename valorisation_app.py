"""
Outil de valorisation par adresse — version Streamlit (locale)
----------------------------------------------------------------
Installation (une seule fois) :
    pip install streamlit pandas openpyxl requests

Lancement :
    streamlit run app.py

Charge le fichier Excel au format du gabarit (onglet COMPARABLES),
géocode via l'API Adresse du gouvernement (api-adresse.data.gouv.fr,
gratuite, sans clé), filtre par rayon et calcule une valeur
basse/cible/haute pondérée.
"""

import io
import math
import re
import time
from datetime import datetime

import folium
import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import live_search
import apis
import auth
import dossiers

st.set_page_config(page_title="Valorisation — Savills", page_icon="🏢", layout="wide")

# ── Phase 1 : authentification obligatoire avant tout accès à l'outil ──
auth.require_login()

# ── Logo Savills embarqué en base64 (aucune dépendance fichier) ──────────
_logo_b64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAb8AAAG/CAIAAABHcU4lAAAQAElEQVR4Aey9C5gcV3Wou6uqHzOj"
    "0YxlSWNZL8t6WJIt2xhj4UBwEp8QDgE7BDDhA5MLXMX3JtgXYjAYIvM4VngcBwdik+Q4PoHEfBxi"
    "cRNCDAS4cXIJIUbG+In1lmXJksVIsjyj0cx0V1fV+feU1BqPuququ6u6q7vXfHtqqqv2Xnvtf629"
    "9qO6e0x3T16SEBACQkAI1ErAVPIjBISAEBACtROQ6Fk7MykhBISAEFAqnugpJIWAEBAC3UZAome3"
    "WVzaKwSEQDwEJHrGw1GkCAEh0G0E0hQ9u429tFcICIF2JiDRs52tJ7oLASHQOgISPVvHXmoWAkKg"
    "nQl0XvRsZ2uI7kJACLQPAYme7WMr0VQICIE0EZDomSZriC5CQAi0DwGJnpVtJVeFgBAQAsEEJHoG"
    "85G7QkAICIHKBCR6VuYiV4WAEBACwQQkegbzaeyulBYCQqBzCUj07FzbSsuEgBBIkoBEzyTpimwh"
    "IAQ6l4BEz/TbVjQUAkIgjQQkeqbRKqKTEBAC6Scg0TP9NhINhYAQSCMBiZ5ptEoSOolMISAE4iUg"
    "0TNeniJNCAiBbiEg0bNbLC3tFAJCIF4CEj3j5dnp0qR9QkAInCIg0fMUCfkrBISAEKiFgETPWmhJ"
    "XiEgBITAKQISPU+RkL/NIyA1CYFOICDRsxOsKG0QAkKg+QQkejafudQoBIRAJxCQ6NkJVuzONkir"
    "hUBrCUj0bC1/qV0ICIF2JSDRs10tJ3oLASHQWgISPVvLX2pvNQGpXwjUS0CiZ73kpJwQEALdTUCi"
    "Z3fbX1ovBIRAvQQketZLTsoJgdME5KwbCUj07EarS5uFgBBonIBEz8YZigQhIAS6kYBEz260urQ5"
    "nQREq/YiINGzvewl2goBIZAWAhI902IJ0UMICIH2IiDRs73sJdoKgTACcr9ZBCR6Nou01CMEhEBn"
    "EZDo2Vn2lNYIASHQLAISPZtFWuoRAu1EQHQNJyDRM5yR5BACQkAInElAoueZTOSKEBACQiCcgETP"
    "cEaSQwgIgfoIdHYpiZ6dbV9pnRAQAkkRkOiZFFmRKwSEQGcTkOjZ2faV1gmB9ieQ1hZI9EyrZUQv"
    "ISAE0k1Aome67SPaCQEhkFYCEj3TahnRSwgIgTgJxC9Lomf8TEWiEBAC3UBAomc3WFnaKASEQPwE"
    "JHrGz1QkCgEh0KkEprdLoud0GnIuBISAEIhKQKJnVFKSTwgIASEwnYBEz+k05FwICAEhEJVA/dEz"
    "ag2STwgIASHQiQQkenaiVaVNQkAIJE9AomfyjKUGISAEOpFAq6NnJzKVNgkBIdANBCR6doOVpY1C"
    "QAjET0CiZ/xMRaIQEALdQKAzomc3WEraKASEQLoISPRMlz1EGyEgBNqFgETPdrGU6CkEhEC6CEj0"
    "PG0PORMCQkAIRCcg0TM6K8kpBISAEDhNQKLnaRZyJgSEgBCITkCiZ3RW0XJKLiEgBLqDgETP7rCz"
    "tFIICIG4CUj0jJuoyBMCQqA7CEj0TKedRSshIATSTkCiZ9otJPoJASGQTgISPdNpF9FKCAiBtBOQ"
    "6Jl2CzWin5QVAkIgOQISPZNjK5KFgBDoZAISPTvZutI2ISAEkiMg0TM5tp0iWdohBIRAJQISPStR"
    "kWtCQAgIgTACEj3DCMl9ISAEhEAlAhI9K1GRa/ETEIlCoNMISPTsNItKe4SAEGgOAYmezeEstQgB"
    "IdBpBCR6dppFO7s90johkB4CEj3TYwvRRAgIgXYiINGznawlugoBIZAeAhI902ML0aRZBKQeIRAH"
    "AYmecVAUGUJACHQfAYme3WdzabEQEAJxEJDoGQdFkdGNBKTN3U5Aome3e4C0XwgIgfoISPSsj5uU"
    "EgJCoNsJSPTsdg+Q9reWgNTevgQkerav7URzISAEWklAomcr6UvdQkAItC8BiZ7tazvRXAicIiB/"
    "W0FAomcrqEudQkAItD8BiZ7tb0NpgRAQAq0gINGzFdSlTiGQRgKiU20EJHrWxktyCwEhIAR8AhI9"
    "fQ5yFAJCQAjURkCiZ228JLcQEALBBLrnrkTP7rG1tFQICIE4CUj0jJOmyBICQqB7CEj07B5bS0uF"
    "QPsQaAdNJXq2g5VERyEgBNJHQKJn+mwiGgkBIdAOBCR6toOVREchIATqIZBsGYmeyfIV6UJACHQq"
    "AYmenWpZaZcQEALJEpDomSxfkS4EhEC7E6imv0TPamTkuhAQAkIgiIBEzyA6ck8ICAEhUI2ARM9q"
    "ZOS6EBACQiCIQG3RM0iS3BMCQkAIdBMBiZ7dZG1pqxAQAvERkOgZH0uRJASEQDcRaEX07Ca+0lYh"
    "IAQ6lYBEz061rLRLCAiBZAlI9EyWr0gXAkKgUwm0b/TsVItIu4SAEGgPAhI928NOoqUQEAJpIyDR"
    "M20WEX2EgBBoDwLdHj3bw0qipRAQAukjINEzfTYRjYSAEGgHAhI928FKoqMQEALpIyDRMw6biAwh"
    "IAS6j4BEz+6zubRYCAiBOAhI9IyDosgQAkKg+whI9EyPzUUTISAE2omARM92spboKgSEQHoISPRM"
    "jy1EEyEgBNqJgETPdrJWFF0ljxAQAs0hINGzOZylFiEgBDqNgETPTrOotEcICIHmEJDo2RzO7VaL"
    "6CsEhEAYAYmeYYTkvhAQAkKgEgGJnpWoyDUhIASEQBgBiZ5hhOR+/QSkpBDoZAISPTvZutI2ISAE"
    "kiMg0TM5tiJZCAiBTiYg0bOTrdsZbZNWCIF0EpDomU67iFZCQAiknYBEz7RbSPQTAkIgnQQkeqbT"
    "LqJV3AREnhCIm4BEz7iJijwhIAS6g4BEz+6ws7RSCAiBuAlI9IybqMjrZALSNiFwmoBEz9Ms5EwI"
    "CAEhEJ2ARM/orCSnEBACQuA0AYmep1nImRBoDgGppTMISPTsDDtKK4SAEGg2AYmezSYu9QkBIdAZ"
    "BCR6doYdpRXdR0Ba3GoCEj1bbQGpXwgIgfYkINGzPe0mWgsBIdBqAhI9W20BqV8ItJKA1F0/AYme"
    "9bOTkkJACHQzAYme3Wx9absQEAL1E5DoWT87KSkEhIBPoDuPEj270+7SaiEgBBolINGzUYJSXggI"
    "ge4kINGzO+0urRYC6SPQbhpJ9Gw3i4m+QkAIpIOARM902EG0EAJCoN0ISPRsN4uJvkJACAQRaN49"
    "iZ7NYy01CQEh0EkEJHp2kjWlLUJACDSPgETP5rGWmoSAEGgXAlH0lOgZhZLkEQJCQAjMJCDRcyYR"
    "eS0EhIAQiEJAomcUSpJHCAgBITCTQHj0nFlCXgsBISAEhIBSEj3FC4SAEBAC9RCQ6FkPNSkjBISA"
    "EGhW9BTSQkAICIHOIiDRs7PsKa0RAkKgWQQkejaLtNQjBIRAZxFor+jZWeylNUJACLQzAYme7Ww9"
    "0V0ICIHWEZDo2Tr2UrMQEALtTKAbo2c720t0FwJCIC0EJHqmxRKihxAQAu1FQKJne9lLtBUCQiAt"
    "BCR61msJKScEhEB3E5Do2d32l9YLASFQLwGJnvWSk3JCQAh0NwGJnq21v9QuBIRAuxKQ6NmulhO9"
    "hYAQaC0BiZ6t5S+1CwEh0K4EJHq2q+Wm6y3nQkAINJ+ARM/mM5cahYAQ6AQCEj07wYrSBiEgBJpP"
    "QKJn85mntUbRSwgIgVoISPSshZbkFQJCQAicIiDR8xQJ+SsEhIAQqIWARM9aaEnecAKSQwh0CwGJ"
    "nt1iaWmnEBAC8RKQ6BkvT5EmBIRAtxCQ6Nktlm6vdoq2QiD9BCR6pt9GoqEQEAJpJCDRM41WEZ2E"
    "gBBIPwGJnum3kWhYLwEpJwSSJCDRM0m6IlsICIHOJSDRs3NtKy0TAkIgSQISPZOkK7I7gYC0QQhU"
    "JiDRszIXuSoEhIAQCCYg0TOYj9wVAkJACFQmINGzMhe5KgTiJSDSOo+ARM/Os6m0SAgIgWYQkOjZ"
    "DMpShxAQAp1HQKJn59lUWtS5BKRlaSIg0TNN1hBdhIAQaB8CEj3bx1aiqRAQAmkiINEzTdYQXYRA"
    "MwhIHfEQkOgZD0eRIgSEQLcRkOjZbRaX9goBIRAPAYme8XAUKUKg2whIeyV6ig8IASEgBOohINGz"
    "HmpSRggIASEg0VN8QAgIgdYRaOeaJXq2s/VEdyEgBFpHQKJn69hLzUJACLQzAYme7Ww90V0ICAFN"
    "oDW/Ej1bw11qFQJCoN0JSPRsdwuK/kJACLSGgETP1nCXWoWAEEgbgVr1kehZKzHJLwSEgBDQBCR6"
    "agryKwSEgBColYBEz1qJSX4hIASEgCZQOXrqO/IrBISAEBAC1QlI9KzORu4IASEgBKoTkOhZnY3c"
    "EQJCQAhUJ5Bk9Kxeq9wRAkJACLQ7AYme7W5B0V8ICIHWEJDo2RruUqsQEALtTiD90bPdCYv+QkAI"
    "dCYBiZ6daVdplRAQAkkTkOiZNGGRLwSEQGcS6Jbo2ZnWk1YJASHQOgISPVvHXmoWAkKgnQlI9Gxn"
    "64nuQkAItI6ARM9a2EteISAEhMApAhI9T5GQv0JACAiBWghI9KyFluQVAkJACJwiINHzFInm/ZWa"
    "hIAQ6AQCEj07wYrSBiEgBJpPQKJn85lLjUJACHQCAYme7WpF0VsICIHWEpDo2Vr+UrsQEALtSkCi"
    "Z7taTvQWAkKgtQQkeraWf6trl/qFgBCol4BEz3rJSTkhIAS6m4BEz+62v7ReCAiBeglI9KyXnJQ7"
    "TUDOhEA3EpDo2Y1WlzYLASHQOAGJno0zFAlCQAh0IwGJnt1o9XS2WbQSAu1FQKJne9lLtBUCQiAt"
    "BCR6psUSoocQEALtRUCiZ3vZS7QNIyD3hUCzCEj0bBZpqUcICIHOIiDRs7PsKa0RAkKgWQQkejaL"
    "tNTTTgREVyEQTkCiZzgjySEEhIAQOJOARM8zmcgVISAEhEA4AYme4YwkhxCoj4CU6mwCEj07277S"
    "OiEgBJIiINEzKbIiVwgIgc4mINGzs+0rrWt/AtKCtBKQ6JlWy4heQkAIpJuARM9020e0EwJCIK0E"
    "JHqm1TKilxCIk4DIip+ARM/4mYpEISAEuoGARM9usLK0UQgIgfgJSPSMn6lIFAKdSkDaNZ2ARM/p"
    "NORcCAgBIRCVgETPqKQknxAQAkJgOgGJntNpyLkQEALJE+iUGiR6doolpR3tTMCwVJQUsYlRRJEn"
    "ojTJVo2ARM9qZOS6EGgegdEnzZFHw5MzqoNssFruuIoiamSrFSxH7oYSkOgZikgyCIEQAszjjKw6"
    "mWoPShR/9s7MtutzO/8gG5xO7DRVRgX8IGrioBEqigz7v2RFicUBdbX6Vuvrl+jZehuIBu1LgGhl"
    "9Hmeoyb2G35yjiuuEElralRpzPDGVOmgEZCcUSOizFKYKOqKKEqyBRCQ6BkAR24JgaoEdNy0VGHY"
    "2H9v9qkNuZ+/Nff4q/Icn3hvft+XskRSP0PV8tNvnOqFRr8KSNNLhJ4HyOFWaHHJEIXAKbtFySt5"
    "hIAQmCJAZGS+uf++zGO/lt+3MXPiQZM5I1GJY2GLsf/2nQGLFAAAEABJREFUDJF0370ZPQ/NThUI"
    "PrjBt2u+W/vmQc1VdEyBRhoi0bMRelK2GwkQOt0Jtf/LmQMbMyyBM6cmjLAggJK4wvnB2zN7vpB1"
    "T4Q/5yGzpHYkINGzHa0mOreUQEYd+EaG4EigJFVUheuko/dae+/JqrxXMY9cbHcCEj3b3YKif1MJ"
    "8Dho9FFz+L7AJ9/TNDp8jzXykGXkp10681R64ZlM2uHKabu1g7aioxBoMQFW4kcetOzt+vFOqCpM"
    "P1naH/qG5fIgPmAzMu59z1DFJEMsBCR6xoJRhHQFAXY87ePG5I6o7xzyoYzvMMf31VbELyjHlBOQ"
    "6JlyA4l66SLA86Lxn9UQCpl+Os+pyYPS0dJlx1i0iduosSglQoRAWgk4k/o97TVpVzpouEWlArpa"
    "wK2aapLMzSUgdmsub6mtzQlYPSqzsLZn6Ew/zZxSAZubAbfaHFdnqy/Rs7PtK62LmYDZq/peXkP0"
    "5KlR7kIvN6+GIjFrLOISI5DO6JlYc0WwEGiAgOeo7Gyv5wKvJhk953l9i5s6vXRq0k8y10tAome9"
    "5KRcVxIwZ6m5VznZ1YpJZUQAc65yMnMVkbdqfrPqnfpuBLw5qj6BUqoigbjtVrESuSgEOoWAZ6vB"
    "K9yhd5WiNIgIO+d6Z8GbnaBNTwQ1dWJKfZLiIdDJ0TMeQiJFCMwgUFKL31FaeFupNFZ1Bkrc5O7g"
    "de7KW21CZ9DEc4Zwedk+BCR6to+tRNN0ECAUGpZa8p7SijtsnqcTJYmVZdU4J/FcfsltpVV/VLRm"
    "K6ar5bty0kkEJHp2kjWlLU0i4AfQc97irPmGveBWl21QYqif0GDoZueCe2zCa2ZO4HYnWf0kvdDn"
    "0G5HsVuYxeS+EKhEgADK5cGLnBU3FS9/cPLVz06+4uHCL22bvPLA5PIP2gOX6r1MPw/ZQpLOG5JF"
    "bqeQgETPFBpFVGobAqzKvYLSyVH5BZ7ZO3XOFbttmiCK1k1Aomfd6KSgEHgJAR1J5Z2WL0HS4S8k"
    "ejbHwFKLEBACnUZAomenWVTa034EpBe2n820xmI3TUF+hUArCbgxVy77BzEDrSJOomcVMKm8LEoJ"
    "gSgE5JOaUSg1nkeiZ+MMRYIQEALdSECiZzdaXdqcLgLSC9Nlj6jaiN2ikuqcfNKStBGIe98zbe3r"
    "VH0kenaqZaVdQkAIJEtAomeyfEW6EBACnUqgnaKnYamZKasMP730Vqdai3ZVJQCHl0Lwc1IkmSRS"
    "4yPQTr2woVb7Pnn6iNP66QzXbaiaZhVOtd00ZeDmddD0isqdUKVjOk0cMMa2G6NPmiOPmCMP6SPn"
    "XJl41igMG85x5Y5rfgYFKW7p8zb91QQspRuS1y0AAq0DAi2lySOPTjUfAiRQPGpyscyBbGSmCCV9"
    "CXqksXglKWUEOnTfU3svHTCvHRjiuCIOiVvSSfFSfPVk/8V7T7kut8hDIrN2V8oiIa1Om8bo6UMH"
    "NxDHthpHfmDtuz+7957s7j/J7vzj3I6NuV0bszs35nZ+MLfzD7I735/luOOG7Pab8ts+lNv9qeyO"
    "T+X2fCH73NcyFBx93MRaiNKWyPK3PZJPgKPnKIaKkYdN2kKLgEDrIEBLTzcfAiRQkGCyMcfdHR/J"
    "brstt+uzGhoFKY6Qif0GIxAIDJwyrR6JepLanYDubpbC2XC5kR+bw9+29n8588zdum/iljjnS/qv"
    "770fzNGF/f6L3+Lqw9+1KDvxjEEvAIgvk5P0pHRFT92r+zziHTNKnzWUd707e/AWa/hO6+i91shm"
    "c/QB88SDZmGLYW/X/1m7dNDwEy+5yF3ykPPAxgwFCTFYa8+dWSxBINby8+mBX0ET7SJ9HsMGcf/Q"
    "/daOj2f1UPEHWdpCi4BA62gjLaW9fsOnH7nILRKIxv7JPPZVDY2CFGeM0aI2ZQE7+qhJFUafB5AK"
    "SkS45OupJSAkNFUftyJUlWCWGlpxxmADukjNT2vbk8Dq80QyfQ3v3c74zUTn/dk9N2QP3p45fLfu"
    "v7glzomL4qszXJcr+DYejt/i6pRibqSnAh/XkyEmqied9gxbUGNLUiqiJ5MsuNP+sZ8be/577vHf"
    "6dn2nhz44Ahlo1/VnbAH1sJsz9yc3fbu3NYP5AgcVER1VMpJShLKoBLK4HYMGE+8N89seu8tWdwI"
    "V8PJ6iZQLogQRCEQsAinClBP7DbKVVN79MSM4NB9mYipcMjwWxddfhNyohK0IzVhc4YRHVBlrVhX"
    "Mp0PLbvvS9lIbY+7Fzb/k5rAhI8zomDy1IYcfQ3vpf/ics6oUXbCWk8oSwTAaZkBbLs+h9OClCkt"
    "dZHK5mjVSdx2q7EdIIA7QwpzQ6A//hv55++0CHmIKYPmvO40XQiWwJxPvjH32NvzVKf7Q17vqNYt"
    "PJaCPoHCsHHo6xaKQYCJc2GLgbZl5TlpvC6ElBPCqQLUP7sq//T7ckf+xWKnGEOgTJSKyOZMqgP3"
    "ZnffEikNf89SZhTBzc1jqgN/m4nShJGfmNbA6W+Jp/ko+vzfWqFl99+emTwaoeVx73s2c3KG2zCW"
    "sFTaeXv2oXU9MCFiMlSXnQ1WjaTpcnBakG5Z07P7jiw7p4ildo6tShFMm5hqrH0IYXRd9vKYpQM9"
    "069I8EqoTiQjHxtQHbuHTB8I3C00AP0QAoe+abETxFiNYqiHkqSECJTFUgV1kVgrsa5np4ndEj2q"
    "R1tp9i31znqtQy9FQnAiz4sPmirS/6Asa1fjSe3ZIe8eV8cfNFEvVP/BV7r46oxKrD4VWnZGkQ57"
    "CUNaRNxkL4jVDKO+TxLXInEr9oRYquDIanL7TXnqZQ2EGqTY64oisDXRk9ayZ8Q6nadAu2/MMiWE"
    "CCmKxo3noSKSHzV4voT5z+wbjdcSLIGQDYQj/2pBYO9NWUYOVCIFl0riLpWSWByxW8JWvTOiQmmw"
    "i2/OVrMvcSkYqhJ5aF1hWG8RhGZuXoaMOr7NLBzUi8rgSrOr1eClrop7ehhcafrv0n+xKQ8kebRw"
    "6G7Ln2w2TW2cihUqiyd2RdlgdY6HO20SurUgehI4mOOwf8HmCHGTVsGCY5MTlZIYMLe+PXfoaxYh"
    "g3DWHB205x0ytv9RtvkjR7UGgsIbU2yJbr81x6gGjWo5T1531eBlbu5Cj1InrwT+OfZvpsp7gVma"
    "ezPvDT9gMXkMrpXW9V/h9J7veYn+p40W9MLgdofcxT2O/HOGPZ8DGzNEMX8+GFIm7tt4LPUWthgs"
    "2rZ+NJrTxq1Ds+0Gd7bqGTF4Btfk8aoiOmzAdTZr9nw+S0xvQgClipF/t578rdyRr+rO6yuADi1P"
    "aEJiSs6aaORhk0EuQCWiSX6BN7A+6pTs2BbLfdEIENjMW5ig9Jwx+u+a/7R6K5zy+GXur7sqm3Dc"
    "dytUnc5LeAW7nPv+Ur+hhciFw5BaqCq1k3gyrKdiYU4bu55NjZ6ETh550066KG0mxd6e+gQyiDHt"
    "YhuFXUi6Vn1CopTC8w59y2KNzMhBpVGKNDkPRmE2sesjObZBmSMH1W6qeb9OeAnKUr43+awxvs+g"
    "75WvtPLEVCd2mt5xj8YGq0F8nXuVowppifvB2iZ9F/MVDhk7N2V5dAM6UtI1RpSPJnQoHsrzNJj+"
    "S4pYsMFszYuedEVmNOyS0E5a26DeLKnKqUFRfnFUYiuaAJrcDBSjPvdP2ec+k0FzqvPrbeSInBmp"
    "EWnlsuimA+jG3MSOwBloSc1e47ItiA7lshVPEFh82hjbbqZk95CJ88gTJn5YUdvyRdo153qHIZ99"
    "3vLFrj0hdJaOqn3/I3Psq1bjAz9sp6fGqeJjCNz3yQzPEhqXFlFCk6In6Ecey+z5dJZuSTsjKlfO"
    "BhdSaUz5ifPMQs9PSOOlf7185AqpXDziCaLYBt39J9kkdugYPIa/Zx28pc79dZpDKjeQE176BKYf"
    "ucitcuIlKWLzp2cDBZbafUeWuQa2m36rfE5MMc/yBl7jRJl/osaJ7SYjU7l4q04Yw5xRdWJbJM+f"
    "c5Vba8RvVbsSrRdoDDnMLegg+EatdWF90nS3REjZb5E2/S7ZeEniek0JmYyIu3kKvcVkzKupbH2Z"
    "I/lQfaLLpUA/8Yyx5+MZf6OkfD30BIKkzCKVX+8NXucuua106fcLr3xs8hWPTV72/YKfXvGQfslF"
    "0st/WFh1j73gRof8zImQT3GO0RMGGL7X2v2FfLWQEV3U9JxIY5d95w1Z5JOm3wo+R3+SNeDRnIE3"
    "ujRtxR32xQ8Uaex0CD4Kjlx8xcMFKMHh3JudWVfruSHFEUIKrmvGXfQc+SfzwP/K0G1m3Dr90jYW"
    "v7d0+mXg2Ys/sAo8eY/2jqhASY3dzOhl++h3TRoYIAhc0OPJWDOiZzN6YUBbQ27Rf8mx584sXSMY"
    "GtmmJxiScD+8F5hDG5zVX7Ff/uOC9tKHpnXhqR6N35LW/bO9dFOJzk4XpiDFSdNlBp+jHvl3vj8b"
    "6clnsKwIdxO3G+iZcTCLGd8S/u6QssIgIEF84W2lVZ8vXvLXhbV/UVj6Prv/Ii8zV5Gs2aqceOkn"
    "no0OvcFZ8cniZf9vYfVdhfM+Y8+53kGOn8rCg09YlRy52zzyLxYhLzhnxLvI4UHZs5/JsokWsQjZ"
    "0JlBmIipCXzRvvCvChd+pUDTFryrNHCZ67e3TKB8wnUe5kBp6NrS8o8VL/5a4cIvF1Z90V60qdR/"
    "jX42gViER0wofPgea/SRqut3Amv+bG/wGjdULG7NZPbEHrMZwSiweV5BTew1YBuYS9+c9TIvN9dj"
    "iq1fJPqrLZNoBY0Jz3v778scuruGBTv+AGEi4PwbnWWfLuG9F99XXPXZ4rzXOr1LdBcueywnOC0J"
    "vyUNXuws2WDT2enyq/7cxvkJAkgjRWwDnsYMdOfGnHNMfwFbxFL1ZUs8eqqMYrOPh2JEpSgqgonE"
    "4LN2c3HlJnvJe3SwALE3anjjBq5Pj9XJUbj1yWQrfcU/FpTOaav+td6CNzkrb7WRM3eDA1PERlGA"
    "POR8/m+tif0GoZ+XjSQksPg98LcZYgc6RBFF7STi5oq77BWf0AQGX+UyMOh2BUDwgZQhwGoqM85K"
    "8cXvKK25vbjmy8XycBJFExRGk4P3Z9wTVbMbOTWP5ypV77/kxrEfmZjsJZea+wJzuOPqyA/D36uE"
    "XgOXuIx8nLQgNVal01jx6aUhMPaEOXxfhqF0+vWAc3yGuLnkDmfVpuL5N9rz/svUW778jsnRd1Hf"
    "Xf2jf8U/kmHKdenyg1e4dH+CwPJ77JpiKH5b2GIc+LsMYSFAz8ZvJRs9QV84aAz/adRaGK+Y5ANr"
    "9e3FwSvd3kX6nSI6YtbuDoAjaRtc6a74kL3mq0UMgPwoyKB/4kHz8HejO0yQ1GMP6W/rQGZQplP3"
    "0JD9IAhcsKm44FqH0Zg7mkC97zcEAsURkpmjQMpwgnBe4uIcQxNqv7DZPL7NrLaRRDzqW6s3FqII"
    "pB9OtvYz7zxt32tOPhLyDJ224C2zLkj5nLCq9eJxXKUwLgPnLyYoEWQAABAASURBVL4VdeyHG2no"
    "ZmfNF4pL324zgzF7T85sVI0/jLK+3zL8D73OWXtnkfUTXQP5ESUdvj8z+oRJCIqYv45sUeNaHaJ1"
    "EVOxccZEmk6oXwb+wmXBjc6lfzc59AaHGY3u9rUHzRk1aBvYCmmMY+vuLS64Va8xqWhGtoovn//T"
    "zOiTZrWoUbHIjIs438QBY+9Nkbb60Ip09nUu25cQINj5ys+QWfdLX5rZpxB+2cMFFvJUF0UaXZHH"
    "fcrSI1nF/P0r3YHXhJsKH2D35sQOs7WfeR/fqr+UC2UqtqV8kWV7z4KE3yRfrizhXliup+YTUx1/"
    "0jx8txWlIO7ElHPd9wvLP2gz79H9l+lkuF+EyNZyHEV3YEW/5it6DkRFIWWU/l4hVnvDD1hEf7qh"
    "SuYnQbsR9dm7HflRpCogwtjCPD8zV/ljTrztRSYQV7wv6ghG72Ia+ItvRvKbqqpm1MGvZfAfpFXN"
    "M3WD5pNn4W0lJt3WoCZAsJu6E/8BFExpV3zUHrxOjyWhFaAY6yCmjRj0zMzoqT+1eWmkaRo0h79h"
    "JfGWBhXxx1XPb86E5mWOM/eXnIrtDS1bT4ZI8OoQ3GgRQs/w9zCaDkbBsnBgZusXfqnYf4mLg+EV"
    "wflrvYtANu7Y0GcSGnHgx2+P3muxbEputI4U2mptqp+fQePYFouORzP8K9WOoJ9/q7vkXSV/nl8t"
    "W4PXtQFsxQ7g0B+6oSpRFxu10GfngbDLy1oTpSaeNY58JdIWG/osvLW0+J3JEig3AdOwIFp208nt"
    "pPL1gBNMWfWBj6v6V7vMO7BjgARu0czRB0w2cIHDyyYnKi0MG8x/USOgalphLVazVnX7e5UYPI7v"
    "tUa+Ex4iIMY2/cpP2ozKxLgAtg3eIi4zu1r18agDP9Ux/cTbOUkihaOpr1Y8le358W1GaHHQM5gs"
    "u77IEEGAC83fYAYUW3yNPe/d+ll8qCimjfoD2rO80JwVMszy2DllAhvcVykIgYHXu4veWjJnqSYQ"
    "oEYSLtW7wlv6+yXayMvQNPIT051Q0KuQ01V9S72+aLuEVMdD/NZMP/OetmaFBrzkEhoygcqwbOfs"
    "JXe67IWpJp5UevbQH9RwvJdnFYs2lPSDTTsoZyz38Fti9JLfs6OM1tRI9J98zmAk4Dz2lFT0JBSO"
    "PWcdm/ood4DSoIcCkyC9XE0ePZoQnghS598YiT6LFiKgw4MOzigcORFl3CPG0e+FTzx9Ahd8oohW"
    "eEbkGmLIyEg++Bpn3vWRBpLCDnN8n4FZ1Rk/GulspT8PfsatMy8A8uh/WonOUM6s1L9CpRGftp/7"
    "tpKywwd+X2wMx8R6YUO6uforXC2lQoXM3+AMXu42zXupiCX8kj8ohc5LyED0Z+BP6JsKErObqyZ2"
    "KgZvGhBMf+Bqb9ZKDyLB2WK8S13mWd7yW8KjNcoXthlju8yKUSNIpantdue58A0jEKGJNUc/mgwS"
    "mNA92zj37agQIh0OpQNqfG91DiU15xUO2RgMQmQpxVa4fvJuhWaMMwOzjxO7DJ62o2SAXPTvWa2a"
    "7JBVt0QCFE34lh7+T6jJ7SEODC5mP4tebze7CSU19yonygoSL3vuTzMqkwivpKIn85ET2/VXzwZr"
    "zZx/3pWlRp5rB8uvdpdpyMDlLps1mL9aHv966aAx+pRJwPVfRjySP+onqd/oDr7KZRoYUXK82dBz"
    "9oUuOyehHMjApmG1ToK5rQUe+w+h6hG8GFRa8OTdVWOPm1gzWENGkvlva4FDBmvVgrsZ/bQ9tF5w"
    "zX2dg+lxgNDMMWagOoLGwreFf84Nf2P6WWLxThyNUYMpUYlFz6KaeDZk7UOH7F3lzlrmtuSLx5mM"
    "zLsW609hCDwUDxleMTDHS2/qcXtcFZ4Pab5faOitTrWQ5GdI+oi2Q9dE+qw6HHgIS/7KKhWMhe+M"
    "tItKCBvZZjW51WzaHt8e4u04JI8Kz3ql22TdKvNs7VXLY/iPosL81ztqIoKrR5FVSx4mHL0XuDzo"
    "x2rB5QibLz5mJTH9DPGnYLWq3aWDFV4w7F8Y1TKUr/dc4E3/jzHl6804mfp+3yj0x582aA7RNrpW"
    "xReMyT0hzcfq1D77olb3VVPll+p/MhHauuIvDHeyeq6SYiDsWx/p+5LHt6iJAwZ+Ul1cnHewHZu2"
    "Y0+Ee3vf+qnhvMlvIQrXK04aEWUx/BN3AjLjwPmFXv5sjxVMQLYEbznGub8baeAf320kMUVLxm5I"
    "LSkWaKHg+s53mYGHZksiA5P/3Fyvb324bLY+7ReM8HzlHKYieo5vMVk1lK9VPKH27OymfJK6YvX+"
    "RVflBiK938ibUI5dlQM8M3MV6zhfavAROMUDVUUFl63nrqs3bQsRvmlh8PWeNbt573w42Za4g7Vz"
    "Um4Df5zAkfKU4J7L63o7yqnijf51FZMP9qmJ48GiJp4x9H/vCB4NgkVUukucq3S54WvOZCQRZl4F"
    "fIglkogGMjElmb08xNOIgKw0CRw1PDgCqq1CLYri1M6jdk5am3IDkd5v5IwrN+xJG89eQ9sCUuCw"
    "MCTghmZuPAMzXJbtI1uwSpAwVMos9M5a69Rg6CB5rbzXeJRgXexG+E5oq0cZuWa29CV1Medl8tF/"
    "RUgXpkzxoOGV+BtzCnGpRmpzRo1GijenbH6O/p7Q0LoKx4wa9sJKivzBJvX7an7Ia3lfJYQZGZU5"
    "K5SBKo0ZIcsfVy/eozyDorKR/zCdUf42IxWGjSjL9lmv9voWt3ojpRk8QupgvCFHxAkQOVuYiN2Z"
    "xeFxBtfVy6a4o13c8k6BZFCyBsJn9ZMjpnLCG39Kavx/rUHPWhw+T6T7MdBFrJ54RP7Q8Z96c2d7"
    "NQTliNXXlQ0vjFKuOB7kMDQ8M0cN8tQlTBbTz9EHzcnmfN2nqdixLURYtvdf6rZmFz4IahjKBO5j"
    "R6TShTkGp9JRVdMD1WBpddwl0M+aGzxROSnVPnbyJMY/CdrNmB0eFu1nlXs8xubUKMpVPUNe9pzw"
    "KF8cqw0UC59QVag3O5CW6Onlwo3ljhhqMmyX1lRz1utvp2ZyHUrgF9/KqN5w+Krhn0P/GDqWqexq"
    "rXlrduHdhlsYtwA4ZAbDTcP+ddw11yjPVJn5KviHoRrXdfDeU/ni+ltbUIhaK96QiZS3sNfgMS4D"
    "SKTcyWSKMsaqMa+mMda2w8FSL0vmZNpUs9RsFptFKGWH5XFV70Kv57zwjkc8O/b3ZtJvdsG13BMq"
    "ymfecovc/tVelGEvDEFH3M+HW5CoVBpTheYsIKpBdVVunlY1eLT2jnv2ifD5QbVKql0P7+TVSgZc"
    "Z+afG/Bwx4A83IL+xE7zxF4zibdiqWg/xC/9REjpxTsGqJiQRIjnGDHpODsGAzYKqyZHqVTtK5Vi"
    "8i2abc1RLIFDWWH9wkFjbKvBNCc0c/0Z8t7xJ01oU12AEDLo9y1GCBkBQjrsFpsYYAlt1OEfWAl9"
    "DjK0ap3BVF7G0CeBXZgHv27RzxXnMZHoiYJmn8ot1GMC5wHJ3q5GHjFbtfdHV0fP8z9sr/t+ITgt"
    "fm8p4s4gjSXnvF9xlm4qrbjDDkgL3uLo2B3FQxHaRsnV3zYfHK3KrQn66qZypkZOHOPof4Z/2wAT"
    "Yb1da5/sh41UWE/ZpHphPbqcLFNSfeeHL0fgdvR7lnPEMCJ9h+1J2TH+4WlE/2InuPP6d89e75A5"
    "xqqVSvKJb+/54dGTxgz/T2vkIbNV9FEAPfsv8vrXVk8Xeb2LIrUFaX4afJW7ZIO94F2lgDTvtQ5x"
    "1s/fwiNrW6+k7OHYVMBHey9w+9aHf3kofe/ItyxnpMpXNzWsEU0rHVbHvk09QbK8MdV/zdT/Pgnd"
    "lwgS08C98DDVgPD6itpGxDefFU79Dwxo11dVg6WYAOnOG9yF1+pP5TRY0ZnFkxr1iIaDl4Q7BTMU"
    "JtUH/iYzsT/hFdyZTT91hd7Obpc+2qrysaCYpZ7KHumvFjhueMGpEElU4plM/d7+4w/G6gmOcU60"
    "Lx9xnlNju8yktm4y6sXHLNY3uJmq/sPsf95VTiuXn9V1a9UdOkLPAo9hh6ElVIfD92eO/tCqtY+E"
    "io2eAW11j6vWf6euR5cWPWesfeal1ebOifTvbvDsYw+YuzZm2QVLdgvsperJqzIB/YaegzX8x9Ny"
    "waonJTV4mZtZGP6pTcZOvXVTVVBjNyxv5KfhHk6MmPOrrioYjVXWaaXpjIPXRZoAMT49+5nsL75t"
    "UaRVM9CW0A/3rXC1KuZwVX7Ii/Lvbiid6VcnHjR3fCQ78rBp9HldZQCa39rEuH3w/gwRJEY1mIZg"
    "/dlXh/c9Kj2xzWR9HbvRWf0U9pvjT4fEROZWc653cnPD3omFou2TnJhUncOUPIIoJkA6gH40u+8v"
    "M+wFEkMjFOqELElFT/oPm3qD6/UEJAonDEAA3XZ9bs9/z5WOKVw/9u4URY1uy8NYxZRhZLMJ/5jb"
    "bqpz3x7tGxyemHrfRczVa3GTe43CNiO0aXOuclvsbHH3wnjGwpKac2WkL0CAtQ95/+2Zp/4gN/Zz"
    "A55+4lYHp7jtNg0V+HjO1ffy8OWbXwgDMBF4/k7ryRvyR/7FKgzrWUP3jGM+hKYdGZ+oa+TfrYj/"
    "8pPMtSVX5c/x+sK+cgmjM20ZfcpkuK1Nflhu5tQjT5jsDARkxN/y673ZayLNkQPkNHqr1fVX1B+L"
    "5Ie8c6+r4fPhLCLH/sn8+Zvz++7NjG033AlF/yUOVJTfARcTjJ7Qz8xVZ7/OiYZJ56IvYYDCFmPX"
    "u7M7/1v2ua9lRh81oU9X1ymeIVVX1LW/GmNeT+0Lh4z9X87sfH8WFGDnGCnp7JEyYn2WwwNXe1Fy"
    "jz1u6jfJRskaLQ8+44wq9gRCsw+sd4kRodm6NIOpBl8ZdfrpI8KXGJOYhO7cmNt7T/bID6yTS8kp"
    "r/PzdMwxweipGbnqnDc4s64Of/OKznzqFwOQGMQObMzs/GBu64dy+/8mM7Z16jumBjxGs1MZ5W8k"
    "AhBjha5TXpWOqpGHzd2fyz79vtzB2zNMzUAdScqpTASmU6chf8k57/JS6LMjFBjZbNrH9XIvRGL0"
    "26aaHDYQG1pi1mrXbP5X0oWqlY4MzN97Ftc2/URxDOrPgYbvtPZ+LLP15tzOT+lHGvp9NVml/bBT"
    "pkHJRk8mIHTd82+1AcqIBNnoiSIklnX0AcLoU7+Rf+wtPTtvzTGaueNaDJ1TJ+zRKcbQrarxVxOw"
    "9Pslp5+UZTChg9XIw+ahzZndm3JP/1+5x34zv/VNucN3W0zwwUsqZ47/xFQ9yzy2bkIlszw59pAZ"
    "5/uWXPXC4xnEBjQQh8yv9/S7Glu+cE62F6pGfvCroWuc/mtqmwBRI+RJzqjB84yj91p43c/fmiOS"
    "7vtStjBs4Jk6zzTX5WXbpcTt5hVU/0XeirvrCaDQxADlRCTFDNvfnX14Xc9TG3L778sc+pY1+rjp"
    "G0OPaXmlV6ZTJqFsJyWc+GRiBXQq0UDnuP6g8cQBY2SrdeRfLYDsuz/CkcrKAAAQAElEQVTLknzH"
    "x7OPvT2/ZU3PQ7B6U273TVki5ugDeh+wzJPiiSZmLvklbn5deCWWUof/wQr5+rtwMadzsON2dHP4"
    "v9Xqv8TtXeKh5+mSLTlrefiu3momQMzN19xezK5WjDfVM1a9U/Y3Fjojm839t2d+ekX+Z1fn99yZ"
    "PXS/dlq8FzfWPRfHnpoM4epVxaXpRuLRk8YSQM++0ll4a/gijswByTcDiwJOGND2bczwxIOl/e5P"
    "ZZ+5O8v0auTHpv8l0pjcN0aAtPTfwoeYuetkKcIBHsY4wUYws+9Df2/t+1p2zz25XZ/N0vxdG7O7"
    "3p9llQSQ/bdYLMmPfVXPLmkjuPwEND9xsZEE2xqKO8b8Xw7/z7EIxKATzxpYjfMGE0LG9xnFsPcq"
    "UcucX3ZVM3oAVbVxov/yAGPV54tM1esLoH7jfffzvZFIynC+9xbttNp7P5vd/zeZ4W/ryRD7pHg7"
    "RbTnZ/mb3tQk3zF71eJ3lIb+z0hvYYlCC0tgBo5MSJlSaUvclOUZyLYP5bbfmtPD2jctAg0LBIOt"
    "0j6PHkUwiiK5tXm0nnmldR7wCJc0gSGBIZrhYetHczs+kt35wdzuG7NEyYO3WIc/axIlaT6hBw54"
    "JEB8LJz4Kebm2DXKK6n+leGf2kRVVtkjj5vxxLKspx9DjSnEVlOXKMCG7Fkvc1r1HQvVFEvndabn"
    "A5e6y2+x656BzmgXpvETTov34sbszu25Qbs3q/udm/T6Cc+f2G/oNSVdmE5hzZDR+pdNip7+hGXp"
    "Bv3FGTguKa6m+zbwj2yyFLYYRBM/mG59e+7R3+rZ+vt5pqWYgWCkY1NW7xKG1N7c24R1nbKKWM+j"
    "cJxm9ydzj/5mD/u8NGHrdTmGaFrEYzT8jCiJdn57zzxyK1UJu7Pum//68FGTrqEX77FoXzCe3xz+"
    "DYmzXu1lFqdg2U6Tm9QLqan+RAAdvNJdfVeh1ofAwVXO8GHcGycnmLJ+wvPZKn302h66A9MI5qSI"
    "0j0FX+EsBampdqMvLXhXae3mom+AGGNomeR0Y3ARY7DVwq7f46/KMyfdd0+GfVLM4NuAI3lalaid"
    "aE7tBZ4ObzHR7elbco9ekX/qutyhqac6KM/d6S3yz7nYTslVg692Qx2ephUPmBO7G128g5R9tELY"
    "N8kz1V34zlJaPp0Z974nrUvCQwigPMNYe2dx/o26hqT7Ly7BzBRT0h2efKN+4Lnr01nWYdiX1mFo"
    "ehAnLUxNjZ600xs3GMFWbrIX3lZiFVAa41qCCQOQ/MUsc1I9oL09t+02/cRp9EmT7RW9txLas+NW"
    "EKtTL0F85CHzua9lWI9ve49+/xCzS19bX2HO4665BfLocr3nef54GVx96YCKYfGe9Y48GGJRun3P"
    "ar2lwI5esEptejek/Q20CmLsgS6/2T7/Ttu3KTAbkBdSlC5AojuQiKQ8NN59S/bp9+TZyDryL9bJ"
    "MJoPEZLc7WZHT1qiu9MSb8l7SuxDD21woE8M5citRBNmIFERQYpNFjYQ2V5ho1rHUDZG4/e4Cq3R"
    "cbPPo0Ye++z849yujTk0YamCVuhGqlAm4UtUTYpaSb37HvN/O3zxjhrHHzfd4/VvrYDXfdFg0zO0"
    "OXPe4Jh9obkkQwUC9F84D73e0XOgOxyctpn9lzDKmmz4Tmv3jVkeN+3/ckZ/MJT+24oY2oLoiUEw"
    "AEf2oVd9wr70h4UFNzps4dN5SFxPNGFsP2EDtleeuTn7s6vz++/NMhNkPohbJFc7aw32Loa/lXn8"
    "d/M89mFLAR18ZTgmV+8MyUCengavc9d+s7jw1hIXZ+SM7WVJDV7qMt0LrWJyj1GY+oRunVWbiqft"
    "k8/qz/hWk4AOONvcX3IwR7U8zb7eml5YfytxY7pw7yJv6dvtK346ueIu238cD1tS/XKjlaSzkMjL"
    "tIPV5M/fnN96U56nBVxpsk1baTcMQOo931vxUfuibxRZy7MWsAb05+KbaQNWBPs2Zh7/nR7mgzqG"
    "WvXPfbBfxeQH5ZGfW1v/nxwPFgtTu3J4AKli/lguwnBGQix42TCB85zrnUWbSuu+X1jzueLg5S4r"
    "Mu4mlOhs+SHvrNfqzbKAKqBBfzi+rX6fxJ1eeDxTfNpAVEBF+TVezzIvRU/b3QBl03sLswLcyKkF"
    "b3Yu+evCyq/YOFV+6psNfMdLWnWsTKIWJiI8Ytr2kRxdjJdNi6H1eypaxpIwAF23d4m39P8uXfhn"
    "xVVf1FuiTIgwgL8i4CSWiqoJwQD+coCHS1tvzh35Vwu3iNEAiCoM6w+Vb/2vWfZeqY5UTZkzr0e5"
    "AqJyApqfmGHhyv3XuPg02/wMTud9xl726dLKzxUv+Jy9+nM2myf9a/X3AbonolRyMg9wTp7V8sfs"
    "VbMvcaM0/PgTej+6Ftkn8zJE6TctbA9/UzdPsfJDHfWVdCcRtOgP/dearea91ln9GXvVpuKyO+y5"
    "G5z8VBjFFX3PTFQ1/IpEDN32Vv3W74lnDNaRidboC2999PT18GMoNuCZ0uJ3lpZ/2GY5yYoAMxAF"
    "MIBvBj9zEkfoE0OZ++z9WIY9aWdEEfUarwgrYsu9X8ywxEAatXBsPAGEBBMSJyDyZ5TgWnGHvfor"
    "9rrNxVV/rl151cftlbfa599oL35nacG1Di4+eIX+Xgw8Xie75q/NJ0jVo7+pZq8Lf+Mnkl/8gVXn"
    "Z95N/bGr0X+3EBKQYKX/60EmIIvcqpkAY6rvTozH57zFWf4B7XvMRpfcViKM4vY4KglfrVl05ALU"
    "gny2RLd9KHfkBxZdr05fjVxjWqKnr7C2ga1YC+SHvMHLXHr7yo/ZL3+woCPpHXpdAB3fBpz4yS8Y"
    "1xEDsJDHADs+lSscMhoMoNhvYreBLdlgRTKpPj39lk4/EiuZUeKaDDBrNxdf9p3Cy787ue7eIrgW"
    "vM2Z92vO4HoXgP2rPUgyJjH1g6rGW1Day0PW0NXVrD3a+rIYHdkm67vQ819WO4KIveDjPzfre9v8"
    "iT0mxRFSTT4MWbbPWuXG+KnQanXVcD1dvbAGxc/MiqG5iMvhe/gh65uXfb1wyfcLOOq5N5+ckGKF"
    "6Yn8cSVMTypsMXimxEYcYhMNoCm1m+7nzslpEd2e6dKCd5VYF6zfNrnun2027JhksXlHEMksApFe"
    "rPn20C8a+4U+6YXN5tPvy4018P9yibwjT1qPX5XHlgisQym/ReWdSlbfy+7S//7ziqcmL/vHydV/"
    "bC99n73gutLgq1xr4LT4MrqXnNQdLk8Lnjqr95m7Lmyqeb+iHw/q88DfYz8y69iUpN8e/f/C/Zll"
    "e2bOSdcK1KKJN9tz3zMAUNn3/Dzsy+Goyz9oE0lf8djk8nvsoZudwev0d9/RhfFw39U5+vkbPNLd"
    "EMWD2T1fzOrPyFgNyqtaPNzbqhZt4g09XRo36B7MoQYvdpZssFd9tsgm6eq7Cis3Fc/9hOMbg0UZ"
    "1MqpEQVZxRP1tt+U1++HyNcsidDJgn3ne/X6EFvWVL6sP45Fu5Z+skQzaeyKTxZxwf6LPAZ2OGgH"
    "hQmpUJP4xjLXO/f0a2XSZy3WQ53/strxha9PfSlkLU7PFGPykDG+I9yfz3qlG+eXOVVrg1yfRoCe"
    "6005KmbKzFVD17I1V1x7Z/HCLxVXfb7IXjyb8nM3nPwqy+mLy2kyajul05Gev9Pa+5dZdyL+58C+"
    "NuHe5udLyVGHDNaeWGLUIIiwycI+6ZK32MtusFf9UXHNV4osY+ff6g680SX0lMNQfcpDn2Xgvj/L"
    "1vrtFbgIq/69d2XZBEBITbXjOowBrMoZny/8qwL7lUOvc2gmjfVGDd8FNYS45pI1KddYZrpQZr4a"
    "/GU3WHeIAeHETrO2GJdRJ3aYDHgUV1V+8AfYDlyc7BsMqlQulzUB7boMwPRfkq3yC7yBy1wW+Ivf"
    "WVrxIXvlptNPO1hZYko8AavpknX9Mgc6fLf1zN1ZlQ/ZMqpLfH3bS/VVFXepk5awtVxW9yzHiDIE"
    "0+U3FNf8SZHQQyRl7sZtbECqwwzYj6fkB7+WoecTExEVKZnq2L+ZPAGkeKT8SqEbGpLYG2IMWLXR"
    "Jmj2nq9NfrKZwSEnYjWNZouh/NB/LYVKYdJ58P5MbR6f9w5/m3JBskE4582umqWpBuVr/r245zC0"
    "tPmNqKNGupVeVjqK/styqneRN3j5yacdzEzXfr24dFOJAY9+QaKPkGqthT5IAB15SD9EqrVsaP64"
    "7RZaYWIZTkYZW1eAMXrP89gQXP7h4pWHJi79fmHBrXqTxd9h0Tki//r0fxHWM8vyeFI08oi595Zs"
    "+UroCT6RX+/xCOiXtk2iMGMA+vvNCS3bXhnoKv0vdwlyNDlAc5if+A+j9JzBBkhAtpfcOmGwVU3B"
    "l1x86QvqXfi2kpoIei/9S0s061Xc+560tFmqx1aPdngedTAzdbRMtvKZlvLQ6RVPTL7ysckVUw+N"
    "ywtKnaOW310fyfH8tgZ3iia8c6LnjPZqY/hr/HGDvcIVN+lNFnZY2KsmZ3DvJcOMdPgf9IdqI00/"
    "S+rAvVG3O1GDxDp91abi0htKDL96bT41AMxQoHNelhQP/UKb44waLz5mRVwa6RFrixksE87s5+TO"
    "9pjvBOeUuykhwFir06jBVikPjVmQ8QAA52GqURrTy7WIejKmsgV38DvZ2E0f4nMR9Ut5Nm2Dgt5k"
    "mfdfnAs+USx/wUFEtaE/+qCpv3tCR8WgQnTj0SfNYw+YFAnKN3WP/kw2dmlX3moT3zEtEX/qTkcf"
    "bGPOev3kneYHtJO7o0+Y7vGALNNuldTR/7RCJ1xn/4pjDU4rJadtQkB3jXGDuQXLsiXvKrEvx1LS"
    "n4dGb8GL/2CO7rAiTYAiC+2K6OnT0DawFY9fhl7nXPA5m7AVfQSjZx7YGBY7qcbyDv8zeTkLSUQH"
    "cpz3GXv5DUVUIr7zslWJjYKmVY0Vehd6PZeHbz5O7jDs40aou5OhdEydeCxoPQ5tBqrei5vWSqko"
    "fgLMLXAejjwPoNfwsJ4nS3ThKDVhfaafoz81KB4lf8Q8XRQ9fSLgI+WHvBUfKKzbXIw4gkGf3ZhD"
    "f28ZfVW7PbsqEzvMF38QPgnyNVl2l73gzQ6dH338K606ekVVOhEUfeJVjAngvN8AZ5BUgPO8Tr9t"
    "PnTMyijW+IVtIZ9tP+tNTv9ip463kQZpGde9ruuFDYEjhlKeXdGL7yvylJWhkcSV0DR8X6Z4NHw8"
    "DpVTztCldiNgscPIY6WVn4v631qYUh5/3KRUmd2ZJyf2mKUDip5/5q3pVxgw59/onPMGR5VS9rbt"
    "6Vomed6/2o0ybk08Z1SdlZfVKynW+KWDIdG/91K97MDu5XIpOon7qVGKmpaYKtoxSur8G/VnZ3gu"
    "HxpA6ZWT29XkXqO2d8IF6t+l0dNnggEGL3dXfNyGrH8l+Dj2hBk0drm6G4dakQz5hd55G2zmqunp"
    "zEw/g9se511X9S31Bl4TMv2kxuMPmc5o0FudmbkXho3xnUGhUwNf7521Nrw6A+swugAAEABJREFU"
    "apTURgToPjiA/w/TonRhS6lD/2jV9k44FfTT1dETMKwCBl6h/98n80FeBiTM4zxXdezCiuy+FfYH"
    "deOy8MV/WOIxIlWXr7TwBM3dSVX8RSTNY9ETp+cJwOxLw79yafxnhn7bfGCtkweMyUdClO+7wJ21"
    "TCZ4gRxP3cQfTp1W/RslT9XCsd7Al1DmnGud2b/qMkyGyj7+oMmCLzRbxAxxRk+aMfmcsePj2dDE"
    "UovMEVVMPJttLHxTKcp397I8HN1ehZipJoeNycDv5aUhGDi/3pvzq26qNuDsUWP8iSrtQukkkqm/"
    "LzmzKOh9JwxXhYPGSKBidJ6RR0zsQuYANfWyfV7Fr6QLKNTEW81lH9AwI6+2/1FI/932kdzuv8gF"
    "PAAIkJ/ELXwgv8Rd9H+EfwqD2nGViT0myz7OG0+x2s1UxXGT6H7kq9axwDQRbY7WePOiSGAOaM1R"
    "izbYUTLbLxjKqvTgyFTOiBH6vbxKqbmvc3Jz09STTUX05IlkcABC8xgTzHtXhC/eWWodvj/jTlSt"
    "mXX92KMhPky75r+6pOyQ+WnVOppwIz3TYsuj5wb33xc2m6MPGjHuHsYAuGDwDCN0Jx1PYPumOGxE"
    "fB+xCvsJ8byw4jPvW6ZrzNbTShQNSBRjxOCYluSqwVdGcmG6a+V/xOgq+4TB1DK0RX0rvLiGvtC6"
    "ImVw1fgzBl4VKXM2Uq6ImYZeF/LPjnAhdvrZ2axIjIvM93k0T7ZqNWKRvvVu73nd9Sb5qNasRI0Y"
    "ZCn95BOqFRN3Vcp+/GAy/22lKA23X4xtHI07evaEc8Wh1WR4tqbmcFV2wGNNrXULrLg0YlT+ziti"
    "0OEQmAjHNXPzKk1dAytN9CaeN/ITM2qXsGPThUd2sy92qRcswUKf/2ZG9VaClvVe+HcruMNwd/5v"
    "c1CJ/qRNOFTrVskcrIS6bnFNLDh4iXan0ApLET+CESpIxTWFnVZTFPp2wdAT1GmlWntKBDF7VP8l"
    "4dPP0lFVmjDOxIYEFeEnt8glTKdn01NP354zxh5upLtFaHaVLOZsNXRzyPQTzUZ+ZLpHKjlMRrGu"
    "J0MV8XpTNdOvN1jTA7yyqiHDbuVCCV3N9IdHT3fEqGyRhHSKItZV1qDHZDl0MNYr91IUieF5YrUb"
    "wSejotAvHjLYzEpVADVyavbqkC9PU0qVxrTmnJyZjIJ35sUZV6w+la4NI1Mx8WR1jNvNULU5L+de"
    "5QSEP3RAscIW4/g2c8aIxfONsZ+ZoZqf9SaHDVa2WRGV3kTHSYlyjhHl+/+94974PmOGRVRLfzBx"
    "70Ivd2FIH7SUso/F9okjM94mZ/u8zNxwkeNP6w/hhedrYg5CeVFFWsA6VbYdWIo2Ud+YqnLV4e9G"
    "/XCUrjLWfU+IzV7jzro6/L0mo0+d4aiuGnv8jItaxdO/rNjn/rqb9olnWd90nMyKMIdwRo3Jg2aq"
    "oifwzL5IUzdyxpVC/K+mali9WgMqvyQk/CNzfItpH0ti2wDZySZm1lZ/5SqYDVW+Me2qM870ddrr"
    "lp4afd7II+aJByN9p8lJTe2Tf+P6Y/aq+a8PWbxTFxNk98Tpt80z1DHXOPLDoLjPCi6/0Jt9kURP"
    "+EVOrsrMVlZYdtiO7zYY/MIyNvW+V9RLw/Aq+43wPNFyxBk9qZEIkj07JHqyFoO+njikZ8Gipj4x"
    "ORqiuW7gLGUy/zpDc92f8+FWYeGvp64xU0evmhMKuy8ah74R2lNqllxbAVP1rfUIc7hEQMHCDlMv"
    "3sufec+oE7sMLuJLAaXmvNnNzk7Tm8Oq6ZoCfyirlj1L94Jgc5D5xDbTHT89nnGltYkdfPu4wYZs"
    "sBosR3L9Ls4fnC3i3bjt5qr8kN67Da3+hf/fSs/WJzRxBRwiVG2rR2UqPv9VKtfjBhenqxefNpyR"
    "8CAbLCeuu8zdjtc08Yyr4ulyXDVrmTv76iB0cLO3Kxbv0yc7PG3n4nRJM84zTDwvcZnbzriexpdB"
    "ra9V38byuyo35LGXEiwFi5z4D2Nsq5mixbupJg8YwS7hNyo76MX17CHu6KnU7DUue7fBYxf0Rx8w"
    "jz9pxtUM1fAPT9JHvxtOI3eOZ82ZmqjOqNHUSx4V9gOWwjGj5T7HQD1xwPjF161S2JdrqIR/2O1h"
    "u2rwUheXCK5qfJvhTH3mnaHOOaYmngkahOCcX+Phii1HHdyotN3FHPmzvZ7levoZrBuec+SRDPsn"
    "wdmaeXdiv37bcqgj9S4Ib11EtcPjRURBJ7O5Kne2x/Tz5MvAPwf+JhPjJ/YDqwq7mVEjj5qlMf0m"
    "4eCs2Tletc8aZWdFmnSPbTOnz6GCq0vq7tTDotp2PJNSRbHb03uxCv7UJpWPPWxNDk9FTFNNHDTG"
    "w/59JiGg294kD6XGk5FTPcsixZcX/8Fk/4SRuPFKY5BgeYf/IWgfvFwFE6C4HiSaZaGxnDB2WYOq"
    "/zI3VBpDBNPP0WT+W1No7TMyEM6OfCscPTr3rfCUM9WHZ4igxbON0BBAoZEfmUFf1ESOhJMx4A1/"
    "zzp4e4bm1FxVNv6tLuAPXuzkL4BgVXVQlUXZmP8lA67ipLDF4GK1AtwaXO+mZWyupmX5esy9sCy3"
    "nhOi4cA6F4DBhcmARYb/ubwVHZw92bsMwKM/tUJnA3pFstCL8R1sCdjNVGdH+PIxH+ezd2YKBw0M"
    "5r9syZFHz0d/aB0L+3caoCc46me4pUpqsn+3wOkLe789Pje+xfjFv2VatWUB6rGfmTtvyFZqQ4Rr"
    "tmKAjJCvxiyWN3BV+Hznhe9Z1M52+Qk/jFapBEuxdzT3VU7lz9RWKdXKy0EDRz16OfUUOlXGVf0r"
    "3YHXh7+NDGd+/k5r5CGT4HWqcAv++js5z3/dilK33mHPh3taFFHkiT96MpXof3mk776FPsPFgf+V"
    "cU+gSWsS0aSw39z7sUwU9gOvqfoFH/Rqa0DpRUFYO6jo8L0WO3dUHZY35vvUOPKIuXNjDh2AX4/0"
    "rMJZ6ykYXMY2zvlV/b9nCXzVMqIwixWm7e6kOnw3LaiWURE7Bta71uxkAn3ValN0I4hOmJrakwfV"
    "nPVQDMuq9Hubdm3MTTzbygkQCh/4bvbFb1p4SLDGNGnBW+McU+OPnroBJTV/Q/ib+HROpV74ujn8"
    "AyuRPulXUP1Ipc6I2vc/MmyBR0HPYjBgmEXarNXhTz+oaHK7OvB3mSY/0NCh8yFzz6ezwWve6rSm"
    "7iQz9+ThQ36Bp+cFU5VUO+D9xx4yT+w1OQFjtWzEjnm/4rRqdl9Nq/a63n9p1AkQ6/e9f5otHDIq"
    "dI3k24xXjz5hMh0JGHd9LcjQs1q/wSOl3+/pa6mPBT2VQFc01i+r/9IHiFzPfjR76FsWK+jqGeO/"
    "Q7DTo9bfZY7eGz5q0RD65Pxfd5hZV1XFVP2rXWvAq5rh1A1EDf9Pa+THZnMcDg+D7fB3rW3vyVUL"
    "nTSQdErBVvw11dxfD/mkLNyOP24OPxC0Q+23YuByN8hSrWhfG9XJYNZ/Ufj3B/otoguPbDb3fjFT"
    "OqzoU/7F5hzx6rGtBhMCIjhqhFZ61m/TPeNckSQy9yQq5Ye8RTfaoe0hg9/s3TdlD23O6JdZDokn"
    "AkrpmNrx8ez+yA9Plt1lBwc77XNr9QTK78ABbaDJjBkseUYeNtEkIGeDt/BmUumoYrXOXieKUfWZ"
    "Mrk+53onv97j5My75Svecc8+kdQajWA39yon068CdED5F75uBY92TEuHbnaCLVVuUVpOEumFjTXO"
    "1f8yCJhRpGCXY1+1Hv+dnsKwdg9cLkqpRvL4vYYnRU/9Rr4Q+PzQrwWnwr0X/qbtF/QvNn5M0G5z"
    "roz0EWa/DXSbvTdln7k7O7FfLwGSM4Dfr9jq3npz7shXLer1FQg4gn7W1e7g5W74nD/rhU6g/Ipw"
    "OAZMhs0j/2KhUuztRSBineOKx+tP3pAfvle3lEr92mccuY7afRdEe3gRLdeMKiK+nHdjyPQTOWjL"
    "sVpifjrvaieut6RUqyXm60kirU9VpgKZ+WrBjU5pLJIAjII/P3V9np7F5Anfi1QsUqaXZPIdm42C"
    "/V/OPP3GHPeommNomnet07PYo12hOaNnSCp6QjA/5J37u05mYfhK1lcXCsN3Wrs2Zg/9vcXEMHYD"
    "IBD0Yz83iNE735/lgVWU0IluNOGctzs9CzwaxcugVDCYQPWFTeJ8CbSXYZMHVrSX58io519v/MgA"
    "i6pHfmDt+mz2mZv1RmdAS/XY8Gsuaps9QTWjLfNltxiUp8F7WGfe5aV8ZIc5szra0n+Nmz/Ha7Po"
    "eWZL0nClpBa+ozRwdfjDd19ZPIQASs+if9HLsCZ+6N+K5agF5hU9Zfjb1u5PZQ/ePrVU7Q+XjVcw"
    "+2FMjVcfKk4qeiKaDkyfHPzNqPQpggEIakxCt92W06vavGJro8E2U1wLGfCw6O6/yG2/KU+MJhBQ"
    "FzWGJtDzNGPeVQ7GC8/sKILg3OuiziXQAU1o785N2YndhtYzH1pJ1QwnW5pXjP9sSuy+Mct6itzU"
    "wjEgLfndkhFhuxYJboFDgim3yOt7ecgGQnD1Ay9z2+Oz7cHNSMFd+m/vEj0BQhd6AcfQhKfhz/Qv"
    "etm+ezNsGeFX9IgofaeacMpqCQOeHze335ZjQjA69f5CqqtWavp1Rymmcb0J/H+BBKMnDaDlKz5k"
    "56PNxchPgghp7J/Mp96Ue+Tqnn335SaeMbypKQ/SdICw9OY052Sekbh4MmVP5qEgxQ/dl3nk1b1s"
    "kRz+rMnwiHzSjLIVX+I0jFrL3l8yZ0XebC6pxdfYg9dFHTPQhESYe/yq/Nab8oR4dNatoAmkqcae"
    "qZvOwK2pDP5dZ1Sxg77vS9mfrO556rocArmOZI4BiXUZ+7mDr3LDNyV8KZOROfj5aznSXZngz74y"
    "6tgzQzbGYpXABgvGmnEr7S+T7YX1t551LhOgRZsqvsO5slhcjkQvY27405f1bP+AngYV2A/FXUl4"
    "LIkTq0Lxk15tqXI3JxPLUCZST70n/59revbckB3ZrGFRBbeiJLzi3JudeW8ssbEeJX9NebQqNRWo"
    "KTP9wZyt1vxJMbs66GnAmTKhw2ITG+y/xXriN/NP35JjOcBzeQLExAEDoAxEp1mD21IEHbb5uIWp"
    "yHbkXy1mmszpiEq7b8kiCpl+OrO6ilfgTm88/1abtSRuVDHPmRd1k2ephW8rURcSzsxQ8QqZSS9s"
    "NgnxjK7778uMPGIS92kOLaXIjMZykWaSYeRR87mvZXZ9Oss2LmX3356hUtAhjVLBiZxnX+cuuNbx"
    "Q2fwyt0XZRcM/yShI7OMgXVurd5SVqbncm/WMrf9lu11jhfldid4QiBb/I7S/MgboL4quJ+feLSw"
    "9U25p9+X231Hdv8pr8Z16arknO7VvMSrcXju+o5NfhybXXsmUkynLKV8mSryDx4+d4Oz7AbbG03E"
    "b5ONnjSTkN97vrdyU80BlLLAIhCAAHYsB1jh/vzN+W0fyO38Yx1MWRpMT3vvye75QnnitZUAAAqW"
    "SURBVJYl/46PZMm2691ZZpr+FAwhiEJg9ESl9OFlny71X+TRhOgFdU5Xzb7YPfcPaxixdSmlfD0Z"
    "XQ9szGif+708zdn9J9k99+Smt5RzxhKa+fTv5clGZh5Ds+NBG30JvrTgIw0ceKO7/MO2MhURn8xG"
    "jkNIKo6ZITkavF3S3zLTu6rOcMKyPTP3ZHMaVESK+wSYNxBAz9tgD17jlqI9QfIL+kffIQtbjMN3"
    "Wzgq7orTsmtJV53h1bzE1enaMxybsr4Q3NuXGfGIh7MEXPxevXCMWKTWbAl3hil1MMDA5e55H7WJ"
    "RzRp6loNB6iVE8WhSXzBGCwNpifCK0GEOEscIVu5CCc1VDaVleLMOlF43q85NYdOpXuv2asW/U6J"
    "VX8dDofCfrK3K5pDoxgGpreUc5pPM8ng5/SPU7pHOqAV2ynnf9jOD3lYxy+TmRX+fK/0XHgeX1p9"
    "R+I46+76Fu+Y7KxX1hl269M2taXY6YtRNzyEMWnFR+sMoGji+6d/xGnZtTzTq3FyLtK1G3Fs6vIT"
    "Hs7kYNkf2r2LTnu4f6uh40sLNyN6+jXOe62z6vP1zED94v7RN0CUo5+/jiOhM7/eW/XnNgrTmeuQ"
    "QBEKWoPqwj8rRvnIAPkrpoSaqR3ranfNF4qsCdCzXHVuKDwyju8w2SEpF0nkxFT6GV1/bbKxGo+b"
    "2nLZTkPj7oUscpEaYyKAEobW3F4c2uCAmlS38CheTZ665aMbiS2ptX9W5KnXdA+vW2a1gnHbrVo9"
    "zMgKauAy98K/KjAmkIsWckxVQiUSE0Y2agevcOuYdU5vDg5nzVEXflm3F7HTb7XqHDVIA1e7hHXt"
    "WNM/zeAYvRG+97D4tOHVvCFRY3NLisejs3816mM3XzqzLWas1jwv0d7i1xX/0Y1fZOwSAatnoB+y"
    "599a/8Z07FrNEIh7swRBw9W3F9lwoA/OyBDvy+ZFT/QmHvWu8FZ/tsieII2kqVxMSUIZRryhm521"
    "dxZRElUbVwwhBKkVn7DZukY+qXGZdUugdhq48LYSDSSsz3QsV3/FCRkC5HOXeSvRk83+gGwN3qKX"
    "qrx39uuifk9Cubp5VzvKTuThQLmKLj/BZ9iSWnFTkU0t1mc4Q3qA4N7og1Y8q1h+Q5H9H7RNWr16"
    "o2e9ehFQrNlq8TtLF9xjMwmlwTS7XmHxlEMB1NBTzq8WeTyXmaNQMh7RzLht/a9Kln/A9t/2QV1x"
    "SY4uh0rLDVzynlLlBhI9w/4hlV/j+HMmD5r886SOBWPeaxwrsnQamF/o9V7gNqHDRFaqMzMyttE7"
    "eB7Azs+CW/X6APgtb6qvw7k3O2iFbujTHE9odvTUDXMUj3cHLnXZmFhy28nvJfPbz91mJiolMQte"
    "cYd90Z8XBy93UQz/iFcHBDJgELbWfrPI2EiNpHirqCaNikjcxbH8BnKOPhzPTEYm/AveKfXiT0xV"
    "5T87cTeWhIbWgPIn7FEEOkot/sNS23yb55lNakEvPFOJGq5gIL2o+v3ihQ+cdmnf02qQ0nBWavRT"
    "/zXu2s3F5f4zIryhYckRBbTMbv7gsPR99su+U1h2h56HWgP6QybgiKh63dmogpRZpAbe6C68w7ns"
    "+4UF73B03Jy+D1i39EoF8TZGbPZSL/lr3VjmuShAqpQ3nmsIJxGs2QO69MeF5R8usubymVesAA0h"
    "0H+FXi9TsFqi7NijZhPiFJsDc35ZbwdW06R8nWl1pl/1Xqya/DZPZ1x/kWhZjYon4IqUXOWO6P/J"
    "U1HI9IuRpKlwxYgwZvgzwqDa8CV8hicZL/t6gchV/pYZtA0qFsc9qiAhKbtaD7HUvu7u4uCr9MoD"
    "lbjetNSy6DnVQuWNGywkCV4XbCqyYTH/RgcioPGTnyeWoy/QPxK8iCkrNxWpdOk7bCaGhLZYagkW"
    "Qi3UpRv7OZsBY3Dq80i+SsEFo9/1pXFkNKYKHn+teN/Uk8dxI9SxjLxa8FsOs1TgVEss1vLrlHtc"
    "f5Qrulb15Zx9kct2RzVNytdR6dxPOLOXOfTn+iqqtZRP8qyrXaouq1HxBJi5AUJjeA3z31YKlUYV"
    "1qAXOkhY/Yp6yRyQqKv/VzxWG6qxH1waAUSuC/6bvWpTEXvheLgfQxpHEnfjSkgjIZn9d4I1Rl/1"
    "+eLKj9nUjkV8TeKqK6KcFkdPtPRbTgxlw+L8G20eUi+/Rz9mYd7kw+JIImcdiYIknzgzTfyJtcbK"
    "TfbyG4qD610qBToK1CG5viLURY35Bd6Ctzmr/qjIsMlzKlqKhiRUrVPsmPKL64HhRgexa24vnvMW"
    "R78nyVYRwwqKDVzuYgLgBCdmqTSkPlUjlkJ+fshb/I5SsCb+3SVvsVEpouRYslFdFN2A2bMg/G0A"
    "NJaNHb8twcdZK71gayKKGqk3WA53gWb26fcmNw4Ez2Gt0L/WW/KuEo7HDtWS20qEUfbEcEtSI45N"
    "WV9Cfr1H0Fxxl33Jdworb7XRf+BSl3qpvfEm1Ceh9dHT1xurk1hdsp8y9DqHIeXCL+ldFWBhBkYb"
    "INaRKAjx1V+xL/r7gp6I/X5x4GK3d5FetFAdya+9yUc6AFUTu4ngy99v09J13ywu3aTfXV9HG1Ge"
    "Nq66x2Zg4Hn6iltsxCKc61TEsU0TiNpU8y5UG2PhbBxxPJ4fLN1QuvCO4kVfL67bXGQiTOCrw7Ep"
    "AkkWo8RiIjJPhJjhLrhWzwlYw3HLr5GTVqW0RE+//dAvE8kPeeyqLLiutO7LhVfunXjlnslLf6h3"
    "WJiZslk5PU0/X3aXTQYmX2S+8sAkBVd/oTjvtQ4DI6FZy3fiGW99hRs5lpWhpYNXuEs22BffX7jy"
    "yMQVT02u+35h5Vds2jK9adPPz2zj0LUlcOG7ZbGN6CZlhUB9BLT7sdxx9JNhlj4sq5d/rHjZdyZx"
    "7Fc8NqkH+M3FAMfGybmL8xMu6QiUuvw/Jng6QgdhXoVKZfmctzylK3pOx6ExFfTGqDdqkJii+8Yg"
    "TCx9V7FaItoOvcHBZmRmSk9BnZAzZdHp8tNzPqOljKv9F3lEfNpSrZlAmNlGdjZpppOeZokm3U5A"
    "z4Twyan+SzfMzNWfl8FvAxwbh+cuzk+4pCNQSifftxN7qFu3ndIbPWc0SYcYIiDGAGVwIg8pfaxn"
    "tKjaS91S9CeFNrNt21it7XK9gwmcDKahjo3bk4eU+qlA20TPxL1KKhACQkAI1EJAomcttCSvEBAC"
    "QuAUAYmep0jIXyEgBIRALQQketZCKzyv5BACQqBbCEj07BZLSzuFgBCIl4BEz3h5ijQhIAS6hYBE"
    "zzRaWnQSAkIg/QQkeqbfRqKhEBACaSQg0TONVhGdhIAQSD8BiZ7pt1G9Gko5ISAEkiQg0TNJuiJb"
    "CAiBziUg0bNzbSstEwJCIEkCEj2TpNsJsqUNQkAIVCYg0bMyF7kqBISAEAgmINEzmI/cFQJCQAhU"
    "JiDRszIXuRovAZEmBDqPgETPzrOptEgICIFmEJDo2QzKUocQEAKdR0CiZ+fZtHNbJC0TAmkiINEz"
    "TdYQXYSAEGgfAhI928dWoqkQEAJpIiDRM03WEF2aQUDqEALxEJDoGQ9HkSIEhEC3EZDo2W0Wl/YK"
    "ASEQDwGJnvFwFCndRkDaKwQkeooPCAEhIATqIfC/AQAA//8wCXf9AAAABklEQVQDAHx/ACWDIo9g"
    "AAAAAElFTkSuQmCC"
)
SAVILLS_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600;700&display=swap');

html, body, [class*="css"] {
  font-family: 'Montserrat','Gotham','Segoe UI',sans-serif !important;
}
.stApp { background-color: #EEE8E3; }

/* ── Sidebar : fond sombre SANS écraser la couleur des popups ── */
[data-testid="stSidebar"] { background-color: #25273A !important; }
/* Texte direct dans la sidebar (labels, captions, headers) */
[data-testid="stSidebar"] > div > div > div > div { color: #EEE8E3 !important; }
[data-testid="stSidebar"] h1,[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 { color: #FFDF00 !important; font-weight:700; }
[data-testid="stSidebar"] label { color: #C0C8D2 !important; font-size:12px; }
/* Les valeurs sélectionnées dans les selectbox (affichées dans la sidebar) */
[data-testid="stSidebar"] [data-testid="stSelectbox"] span { color: #EEE8E3 !important; }
[data-testid="stSidebar"] [data-testid="stNumberInput"] input { color: #EEE8E3 !important; }
[data-testid="stSidebar"] [data-testid="stTextInput"] input { color: #EEE8E3 !important; }
/* Les options du popup dropdown restent avec le style par défaut (fond blanc, texte sombre) */

/* ── Métriques ── */
[data-testid="stMetric"] {
  background:#FFFFFF; border-radius:8px;
  padding:12px 16px; border-left:3px solid #008493;
}
[data-testid="stMetricLabel"] { font-size:12px !important; color:#79828C !important; font-weight:600; text-transform:uppercase; }
[data-testid="stMetricValue"] { font-size:22px !important; font-weight:700 !important; color:#25273A !important; }

/* ── Boutons ── */
.stButton > button {
  background-color:#008493 !important; color:white !important;
  font-weight:700 !important; border:none !important; border-radius:6px !important;
  font-size:14px !important; padding:8px 20px !important;
}
.stButton > button:hover { background-color:#006d7a !important; }
.stDownloadButton > button {
  background-color:#FFDF00 !important; color:#25273A !important;
  font-weight:700 !important; border:none !important; border-radius:6px !important;
}

/* ── Tabs ── */
[data-testid="stTabs"] [data-baseweb="tab-list"] {
  background:#FFFFFF; border-radius:8px 8px 0 0; border-bottom:2px solid #008493;
}
[data-testid="stTabs"] [data-baseweb="tab"] { font-weight:600 !important; color:#79828C !important; }
[data-testid="stTabs"] [aria-selected="true"] { color:#008493 !important; }
[data-testid="stTabsContent"] > div { padding-top:0 !important; }

/* ── Tables ── */
[data-testid="stDataFrame"] th { background-color:#25273A !important; color:white !important; font-weight:700 !important; font-size:13px !important; }
[data-testid="stDataFrame"] td { font-size:13px !important; color:#25273A !important; }

input[type=range] { accent-color:#008493; }
[data-testid="stProgress"] > div { background-color:#008493 !important; }
</style>
"""
st.markdown(SAVILLS_CSS, unsafe_allow_html=True)

# ── Header Savills : logo + titre côte à côte via st.columns (fiable dans tous les contextes) ──
import base64 as _b64_mod
_logo_bytes = _b64_mod.b64decode(_logo_b64)
_hcol1, _hcol2 = st.columns([6, 1])
with _hcol1:
    st.markdown(
        "<div style='background:#25273A;padding:14px 20px;border-radius:8px;"
        "border-bottom:3px solid #FFDF00;margin-bottom:12px'>"
        "<p style='color:white;font-size:26px;font-weight:700;margin:0;font-family:Montserrat,sans-serif'>"
        "🏢 Outil de Valorisation Tertiaire</p>"
        "<p style='color:#79828C;font-size:13px;margin:4px 0 0'>Bureaux · Locaux d'activités · Commerce — Nantes Métropole</p>"
        "</div>",
        unsafe_allow_html=True
    )
with _hcol2:
    st.image(_logo_bytes, width=120)

BAN_URL = "https://api-adresse.data.gouv.fr/search/"


# ---------------------------------------------------------------- géocodage
@st.cache_data(show_spinner=False)
def geocode(query: str):
    """Retourne un dict {lat, lon, label, city, postcode} ou None. Mis en cache."""
    if not query or not str(query).strip():
        return None
    try:
        r = requests.get(BAN_URL, params={"q": query, "limit": 1}, timeout=8)
        r.raise_for_status()
        feats = r.json().get("features", [])
        if not feats:
            return None
        f = feats[0]
        lon, lat = f["geometry"]["coordinates"]
        props = f.get("properties", {})
        return {
            "lat": lat, "lon": lon,
            "label": props.get("label", query),
            "city": props.get("city", ""),
            "postcode": props.get("postcode", ""),
        }
    except requests.RequestException:
        return None


def haversine_m(lat1, lon1, lat2, lon2):
    R = 6371000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlmb / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def to_num(v):
    try:
        if v is None or v == "":
            return None
        return float(str(v).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return None


# ------------------------------------------------- extraction type de bail / preneur
BAIL_RE = re.compile(r"bail\s+(commercial|d[ée]rogatoire|professionnel)\s*(\d/\d/\d)?", re.I)
PRENEUR_RE = re.compile(r"lou[ée]\s+(?:à|a)\s+(?:la\s+soci[ée]t[ée]\s+)?([A-ZÀ-Ü][\w\-' ]{2,40})", re.I)


def extract_type_bail(row_):
    obs = str(row_.get("Observations", "") or "")
    m = BAIL_RE.search(obs)
    if m:
        return f"Bail {m.group(1)}{' ' + m.group(2) if m.group(2) else ''}".strip()
    opv = row_.get("Operation")
    if pd.notna(opv) and opv not in ("", "Non disponible"):
        return str(opv)
    return "Non disponible"


def extract_preneur_etat(row_):
    obs = str(row_.get("Observations", "") or "")
    m = PRENEUR_RE.search(obs)
    if m:
        return m.group(1).strip()
    etat = row_.get("Etat")
    if pd.notna(etat) and etat not in ("", "Non disponible"):
        return str(etat)
    return "Non disponible"


def auto_comment(r: dict) -> str:
    """Génère un commentaire synthétique à partir des champs disponibles.
    Le commentaire est pré-rempli mais reste éditable dans le tableau."""
    parts = []
    # Type d'opération + état
    op  = str(r.get("Operation", "") or r.get("Date_transaction", "") or "")
    etat = str(r.get("Etat", "") or "").strip()
    if "Offre" in op or "offre" in op:
        parts.append("Offre en cours")
    if etat and etat not in ("Non disponible", ""):
        parts.append(etat)
    # Surface
    surf = r.get("Surface_m2") or r.get("Surface (m²)")
    if surf:
        try: parts.append(f"{int(float(surf))} m²")
        except: pass
    # Loyer / Prix
    loyer = to_num(r.get("Loyer_HT_HC_eur_m2_an") or r.get("Loyer HT-HC €/m²/an")
                   or r.get("Loyer_facial_eur_m2_an") or r.get("Loyer facial €/m²/an"))
    prix  = to_num(r.get("Prix_vente_eur_m2") or r.get("Prix vente €/m²"))
    if loyer: parts.append(f"Loyer : {loyer:.0f} €/m²/an HT-HC")
    elif prix: parts.append(f"Prix : {prix:.0f} €/m²")
    # Bail dans observations
    obs = str(r.get("Observations", "") or "")[:400]
    bail_m = re.search(r"bail\s+[\d/]+\s*ans?", obs, re.I)
    if bail_m: parts.append(bail_m.group(0).title())
    # Prestations notables
    if re.search(r"climatisation", obs, re.I): parts.append("Climatisé")
    if re.search(r"fibre", obs, re.I): parts.append("Fibre")
    # Source
    src = str(r.get("Source", "") or "").split(".")[0]
    if src and src != "Non disponible": parts.append(f"Source : {src}")
    return ". ".join(parts) + ("." if parts else "")


def build_export_table(res_df: pd.DataFrame, operation: str = "Location") -> pd.DataFrame:
    """Format professionnel Savills.
    Conditionnel : colonnes loyer/charges pour Location, prix/rendement pour Vente.
    Structure : Identification | Valorisation | Qualificatifs | Commentaire
              | Annexes | Observations | Lien
    """
    rows = []
    is_vente = operation == "Vente"

    for _, r in res_df.iterrows():
        r = r.to_dict()

        # ── Identification ───────────────────────────────────────────────
        date_val = r.get("Date_transaction") or r.get("Date", "")
        if not date_val or str(date_val).strip() in ("", "nan", "Non disponible"):
            date_val = "Offre en cours"

        adresse = str(r.get("Adresse", "") or "").strip()
        commune = str(r.get("Commune", "") or "").strip()
        cp      = str(r.get("Code_postal", "") or "").strip()
        if commune and commune not in adresse:
            adresse_complete = f"{adresse}, {cp} {commune}".strip(", ")
        else:
            adresse_complete = adresse or "—"

        surface = to_num(r.get("Surface_m2") or r.get("Surface (m²)"))
        etat    = str(r.get("Etat", "") or r.get("État", "") or "—").strip()
        type_a  = str(r.get("Type_actif", "") or r.get("Type", "") or "—").strip()
        photo   = str(r.get("Photo_url", "") or r.get("Photo", "") or "")
        lien    = str(r.get("Lien", "") or "")

        # ── Valorisation (conditionnel) ──────────────────────────────────
        if is_vente:
            prix_m2     = to_num(r.get("Prix_vente_eur_m2") or r.get("Prix vente €/m²"))
            prix_total  = to_num(r.get("Prix_vente_total_eur"))
            rendement   = to_num(r.get("Taux_rendement_pct"))
            valeur_cle  = prix_m2
        else:
            loyer_hthc  = to_num(r.get("Loyer_HT_HC_eur_m2_an") or r.get("Loyer HT-HC €/m²/an"))
            loyer_facial= to_num(r.get("Loyer_facial_eur_m2_an") or r.get("Loyer facial €/m²/an"))
            charges     = to_num(r.get("Charges_eur_m2_an") or r.get("Charges €/m²/an"))
            valeur_cle  = loyer_hthc or loyer_facial

        # ── Commentaire (édité ou auto) ──────────────────────────────────
        commentaire = str(r.get("Commentaire", "") or "").strip()
        if not commentaire:
            commentaire = auto_comment(r)

        # ── Annexes ──────────────────────────────────────────────────────
        parkings  = r.get("Nb_parkings") or r.get("Parkings")
        clim      = str(r.get("Climatisation", "") or "")
        fibre     = str(r.get("Fibre", "") or "")
        ascenseur = str(r.get("Ascenseur", "") or "")
        obs       = str(r.get("Observations", "") or "")[:300]

        # Build row
        row = {
            "Date": date_val,
            "Photo": photo,
            "Adresse complète": adresse_complete,
            "Surface (m²)": surface,
            "État": etat,
            "Type d'actif": type_a,
        }
        if is_vente:
            row["Prix de vente €/m²"] = prix_m2
            if prix_total: row["Prix total €"] = prix_total
            if rendement:  row["Rendement %"] = rendement
        else:
            row["Loyer HT-HC €/m²/an"] = valeur_cle
            if loyer_facial and loyer_facial != valeur_cle:
                row["Loyer facial €/m²/an"] = loyer_facial
            if charges: row["Charges €/m²/an"] = charges

        row["Commentaire"] = commentaire
        # Annexes
        if parkings: row["Parkings"]   = parkings
        if clim and clim not in ("Non disponible", ""):     row["Clim"] = clim
        if fibre and fibre not in ("Non disponible", ""):   row["Fibre"] = fibre
        if ascenseur and ascenseur not in ("Non disponible", ""): row["Ascenseur"] = ascenseur
        if obs: row["Observations"] = obs
        row["Lien"] = lien
        rows.append(row)

    return pd.DataFrame(rows)


def _download_photo(url: str, timeout: int = 6) -> bytes | None:
    """Télécharge une photo pour l'intégrer dans l'Excel. Retourne None en cas d'échec."""
    if not url or not url.startswith("http"):
        return None
    try:
        r = requests.get(url, timeout=timeout, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code == 200 and "image" in r.headers.get("Content-Type", ""):
            return r.content
    except Exception:
        pass
    return None


def export_to_excel_bytes(export_df: pd.DataFrame, title: str = "Extraction comparables",
                          operation: str = "Location") -> bytes:
    """Export Excel professionnel charte Savills avec sections visuelles."""
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter as gcl
    import tempfile, os

    C_DARK  = "25273A"; C_TEAL  = "008493"
    C_LIGHT = "EEE8E3"; C_WHITE = "FFFFFF"; C_LINK  = "008493"

    SECTION_COLS = {
        "Identification": {"Date":14,"Photo":22,"Adresse complète":42,"Surface (m²)":12,"État":14,"Type d'actif":14},
        "Valorisation":   {"Loyer HT-HC €/m²/an":18,"Loyer facial €/m²/an":18,"Charges €/m²/an":14,
                           "Prix de vente €/m²":16,"Prix total €":16,"Rendement %":12},
        "Commentaire":    {"Commentaire":50},
        "Annexes":        {"Parkings":10,"Clim":10,"Fibre":10,"Ascenseur":12},
        "Observations":   {"Observations":52},
        "Source":         {"Lien":28},
    }

    available = set(export_df.columns)
    ordered_cols, section_ranges = [], {}
    for sect, cols in SECTION_COLS.items():
        present = [c for c in cols if c in available]
        if present:
            s = len(ordered_cols) + 1
            ordered_cols.extend(present)
            section_ranges[sect] = (s, len(ordered_cols))

    n_cols = len(ordered_cols)
    thin  = Side(style="thin", color="D9D9D9")
    BRD   = Border(left=thin, right=thin, top=thin, bottom=thin)
    CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    TOPL  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    wb = Workbook(); ws = wb.active; ws.title = "Comparables"
    ws.sheet_view.showGridLines = False

    # Ligne 1 : titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    t = ws.cell(1,1,f"  {title}")
    t.font = Font(name="Calibri",bold=True,size=14,color=C_WHITE)
    t.fill = PatternFill("solid",start_color=C_DARK); t.alignment = LEFT
    ws.row_dimensions[1].height = 32

    # Ligne 2 : en-têtes de section
    for sect,(c1,c2) in section_ranges.items():
        ws.merge_cells(start_row=2,start_column=c1,end_row=2,end_column=c2)
        c = ws.cell(2,c1,sect.upper())
        c.font = Font(name="Calibri",bold=True,size=9,color=C_WHITE)
        c.fill = PatternFill("solid",start_color=C_TEAL)
        c.alignment = CTR; c.border = BRD
    ws.row_dimensions[2].height = 18

    # Ligne 3 : en-têtes colonnes
    all_widths = {}
    for d in SECTION_COLS.values(): all_widths.update(d)
    for j,col in enumerate(ordered_cols,start=1):
        ws.column_dimensions[gcl(j)].width = all_widths.get(col,14)
        c = ws.cell(3,j,col)
        c.font = Font(name="Calibri",bold=True,size=10,color=C_WHITE)
        c.fill = PatternFill("solid",start_color=C_DARK)
        c.alignment = CTR; c.border = BRD
    ws.row_dimensions[3].height = 28

    photo_j = (ordered_cols.index("Photo")+1) if "Photo" in ordered_cols else None
    lien_j  = (ordered_cols.index("Lien") +1) if "Lien"  in ordered_cols else None
    LEFT_C  = {"Adresse complète","Commentaire","Observations"}
    FMT_EUR = {"Loyer HT-HC €/m²/an","Loyer facial €/m²/an","Charges €/m²/an","Prix de vente €/m²","Prix total €"}
    tmp_files = []

    for i,row_ in export_df.iterrows():
        r = i + 4
        bg = PatternFill("solid",start_color=C_LIGHT) if i%2==0 else None
        for j,col in enumerate(ordered_cols,start=1):
            val = row_.get(col,"")
            c = ws.cell(r,j)
            c.border = BRD; c.font = Font(name="Calibri",size=10)
            if bg: c.fill = bg
            if j == lien_j and val and str(val).startswith("http"):
                c.hyperlink = str(val); c.value = "🔗 Voir"
                c.font = Font(name="Calibri",size=10,color=C_LINK,underline="single")
                c.alignment = CTR; continue
            if j == photo_j:
                c.value = ""; c.alignment = CTR
            elif val is None or (isinstance(val,float) and math.isnan(val)):
                c.value = ""; c.alignment = CTR
            else:
                c.value = val
                c.alignment = TOPL if col in LEFT_C else CTR
            if col == "Surface (m²)" and isinstance(val,(int,float)): c.number_format="#,##0"
            elif col in FMT_EUR and isinstance(val,(int,float)): c.number_format="#,##0 €"
        ws.row_dimensions[r].height = 75
        if photo_j:
            img_bytes = _download_photo(str(row_.get("Photo","") or ""))
            if img_bytes:
                try:
                    with tempfile.NamedTemporaryFile(suffix=".jpg",delete=False) as tmp:
                        tmp.write(img_bytes); tp = tmp.name
                    xi = XLImage(tp); xi.width,xi.height = 110,70
                    ws.add_image(xi,f"{gcl(photo_j)}{r}"); tmp_files.append(tp)
                except Exception: pass

    ws.freeze_panes = "A4"

    # Onglet Synthèse
    ws2 = wb.create_sheet("Synthèse"); ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 30; ws2.column_dimensions["B"].width = 24
    ws2.merge_cells("A1:B1")
    h = ws2.cell(1,1,f"  Synthèse — {title}")
    h.font = Font(name="Calibri",bold=True,size=14,color=C_WHITE)
    h.fill = PatternFill("solid",start_color=C_DARK); h.alignment = LEFT
    ws2.row_dimensions[1].height = 32
    val_col = "Loyer HT-HC €/m²/an" if operation != "Vente" else "Prix de vente €/m²"
    if val_col in export_df.columns:
        vals = export_df[val_col].apply(to_num).dropna()
        unit = "€/m²/an" if operation != "Vente" else "€/m²"
        for idx,(label,val) in enumerate([
            ("Comparables analysés", len(vals)),
            ("Valeur basse", f"{vals.min():,.0f} {unit}" if len(vals) else "—"),
            ("Médiane",      f"{vals.median():,.0f} {unit}" if len(vals) else "—"),
            ("Valeur haute", f"{vals.max():,.0f} {unit}" if len(vals) else "—"),
            ("Moyenne",      f"{vals.mean():,.0f} {unit}" if len(vals) else "—"),
        ], start=2):
            a = ws2.cell(idx,1,label); a.font = Font(name="Calibri",size=11,bold=True)
            b = ws2.cell(idx,2,val);   b.font = Font(name="Calibri",size=11,color=C_TEAL)
            ws2.row_dimensions[idx].height = 20

    buf = io.BytesIO(); wb.save(buf)
    for p in tmp_files:
        try: os.unlink(p)
        except: pass
    return buf.getvalue()


# ── Suppression du titre Streamlit par défaut (remplacé par le header Savills) ──


auth.render_user_badge()

# ══════════════════════════════════════════════════════════════════════════
# PHASE 2 — Gate dossier : l'utilisateur doit ouvrir ou créer un dossier
# ══════════════════════════════════════════════════════════════════════════
user = auth.current_user()

if "active_dossier_id" not in st.session_state:
    st.session_state.active_dossier_id = None

if st.session_state.active_dossier_id is None:
    st.markdown(
        "<div style='max-width:640px;margin:20px auto'>"
        "<h2 style='color:#25273A'>📁 Mes dossiers</h2></div>",
        unsafe_allow_html=True,
    )

    existing = dossiers.list_dossiers(user["id"]) if user else []

    col_new, col_list = st.columns([1, 2])
    with col_new:
        st.markdown("#### ➕ Nouveau dossier")
        with st.form("new_dossier_form"):
            nc = st.text_input("Nom du client / dossier", placeholder="ex. Actif Rue de la Tour")
            adr = st.text_input("Adresse (optionnel, éditable ensuite)", placeholder="ex. Nantes")
            tb = st.selectbox("Type de bien", ["Bureaux", "Activités", "Commerce", "Mixte"])
            create = st.form_submit_button("Créer le dossier", use_container_width=True)
        if create:
            new_d = dossiers.create_dossier(user["id"], nc, adr, tb)
            if new_d:
                st.session_state.active_dossier_id = new_d["id"]
                st.session_state.active_dossier_meta = new_d
                st.rerun()
            else:
                st.error("Échec de création — vérifie la connexion à la base de données.")

    with col_list:
        st.markdown("#### 📂 Dossiers existants")
        if not existing:
            st.caption("Aucun dossier pour l'instant. Crée ton premier dossier à gauche.")
        else:
            for d in existing:
                with st.container(border=True):
                    c1, c2, c3 = st.columns([3, 1, 1])
                    with c1:
                        st.markdown(f"**{d.get('nom_client', 'Sans nom')}**")
                        st.caption(f"{d.get('adresse','—') or '—'} · {d.get('type_bien','—')} · "
                                   f"{d.get('statut','brouillon')}")
                        upd = d.get("updated_at", "")
                        if upd:
                            st.caption(f"Modifié le {upd[:10]} à {upd[11:16]}")
                    with c2:
                        if st.button("Ouvrir", key=f"open_{d['id']}", use_container_width=True):
                            st.session_state.active_dossier_id = d["id"]
                            st.session_state.active_dossier_meta = d
                            st.rerun()
                    with c3:
                        if st.button("🗑️", key=f"del_{d['id']}", use_container_width=True):
                            dossiers.delete_dossier(d["id"])
                            st.rerun()
    st.stop()

# ── Dossier actif : chargement du contenu sauvegardé (une seule fois) ──────
if "dossier_loaded" not in st.session_state:
    full = dossiers.load_dossier(st.session_state.active_dossier_id)
    if full:
        st.session_state.active_dossier_meta = full
        dossiers.restore_session_state(full.get("data") or {}, st.session_state)
    st.session_state.dossier_loaded = True

# ── Bandeau dossier actif (en haut, au-dessus de tout le reste) ────────────
_meta = st.session_state.get("active_dossier_meta", {})
_bcol1, _bcol2, _bcol3 = st.columns([5, 1, 1])
with _bcol1:
    st.markdown(
        f"<div style='background:#FFF9E5;border-left:3px solid #FFDF00;"
        f"padding:8px 14px;border-radius:6px;font-size:13px;color:#25273A'>"
        f"📁 Dossier actif : <b>{_meta.get('nom_client','Sans nom')}</b> — "
        f"{_meta.get('adresse','—') or '—'}</div>",
        unsafe_allow_html=True,
    )
with _bcol2:
    if st.button("💾 Sauvegarder", use_container_width=True):
        ok = dossiers.save_dossier_data(
            st.session_state.active_dossier_id,
            dossiers.snapshot_session_state(st.session_state),
        )
        st.toast("✅ Sauvegardé" if ok else "❌ Échec de la sauvegarde", icon="💾")
with _bcol3:
    if st.button("📂 Changer", use_container_width=True):
        st.session_state.active_dossier_id = None
        st.session_state.dossier_loaded = False
        st.rerun()

with st.sidebar:
    st.header("1 — Actif à valoriser")
    _restored_addr = st.session_state.get("address_value", "")
    address = st.text_input("Adresse (recommandé pour la précision)",
                             value=_restored_addr,
                             placeholder="ex. 5 rue de la Tour, 44200 Nantes")
    st.session_state["address_value"] = address
    col_c1, col_c2 = st.columns(2)
    with col_c1:
        manual_commune = st.text_input("Commune", value="", placeholder="ex. Nantes")
    with col_c2:
        manual_cp = st.text_input("Code postal", value="", placeholder="ex. 44000")

    if address:
        target = geocode(address)
        if not target:
            st.error("Adresse introuvable ou réseau indisponible. Vérifie l'orthographe / ajoute la commune.")
    elif manual_commune.strip():
        target = geocode(f"{manual_commune.strip()} {manual_cp.strip()}".strip())
        if target:
            st.caption("ℹ️ Pas d'adresse précise saisie : centrage sur la commune (moins précis).")
        else:
            st.error("Commune introuvable. Vérifie l'orthographe.")
    else:
        target = None

    if target:
        st.success(f"✓ {target['label']}")
        # mini-carte satellite Leaflet/Esri — pas de clé API nécessaire
        lat, lon = target["lat"], target["lon"]
        mini_map_html = f"""
        <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
        <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
        <div id="minimap" style="height:200px;border-radius:8px;overflow:hidden"></div>
        <script>
          var map = L.map('minimap', {{zoomControl:false, attributionControl:false}})
                     .setView([{lat},{lon}], 17);
          L.tileLayer('https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{{z}}/{{y}}/{{x}}',
            {{maxZoom:19}}).addTo(map);
          L.marker([{lat},{lon}]).addTo(map)
           .bindPopup("<b>{target['label']}</b>").openPopup();
        </script>
        """
        components.html(mini_map_html, height=210)
        st.caption("Vérifie que le point correspond bien au bien recherché.")

    st.header("2 — Critères")
    st.caption("Ces critères pilotent à la fois la recherche en direct et le filtrage des résultats — "
                "un seul réglage, pas de doublon.")
    op = st.selectbox("Opération", ["Location", "Vente", "Tous (Location + Vente)"])
    asset_type = st.selectbox("Type d'actif", ["Bureaux", "Activités", "Commerce", "Tous"])
    surface_target = st.number_input("Surface de l'actif (m²) — optionnel", min_value=0, value=0, step=10)
    radius_km = st.slider("Rayon de recherche (km)", 0.2, 15.0, 5.0, 0.1)

    st.subheader("Pondération du score")
    w_prox = st.slider("Proximité", 0, 100, 45)
    w_sim = st.slider("Similarité (surface)", 0, 100, 35)
    w_rec = st.slider("Récence", 0, 100, 20)

    st.header("3 — Source des comparables")
    mode = st.radio("Mode", ["Charger un fichier Excel", "Rechercher en direct (expérimental)"])

    file = None
    selected_sources = []
    if mode == "Charger un fichier Excel":
        file = st.file_uploader("Fichier Excel (gabarit, onglet COMPARABLES)", type=["xlsx"])
    else:
        if not target:
            st.warning("Renseigne d'abord l'adresse ou la commune ci-dessus : la recherche en "
                        "direct s'appuie dessus.")
        else:
            st.caption(f"📍 Recherche centrée sur **{target['city'] or target['label']}** "
                        f"({target['postcode'] or '?'}).")
        st.caption(f"🔎 Recherchera : **{op}** / **{asset_type}** (réglages de la section 2 ci-dessus).")
        selected_sources = st.multiselect(
            "Sources à interroger",
            options=list(live_search.ALL_SOURCES.keys()),
            default=["BureauxLocaux", "Geolocaux", "ArthurLoyd", "TournyMeyer (JLL)"],
            help="CBRE est marqué « non vérifié » : l'URL de recherche a été extrapolée à partir "
                 "d'un seul exemple observé, pas confirmée pour tous les cas.",
        )
        max_pages = st.slider(
            "Pages à scraper par source",
            min_value=1, max_value=10, value=3,
            help="1 page ≈ 30 annonces (BureauxLocaux). "
                 "3 pages = ~90 annonces. Plus de pages = plus long (~0.4s/page).",
        )
        only_complete = st.checkbox(
            "Ne garder que les annonces complètes (adresse précise + prix exploitable)",
            value=True,
        )
        run_live = st.button("Lancer la recherche en direct", disabled=not target)

    # ── Section APIs ──────────────────────────────────────────────────
    st.header("4 — APIs externes")
    with st.expander("🎯 Targomo (isochrones)", expanded=False):
        import os as _os
        _env_key = _os.environ.get("TARGOMO_API_KEY", "")
        targomo_key = st.text_input(
            "Clé API Targomo",
            value=_env_key,
            type="password",
            placeholder="ta-clé-ici",
            help="Gratuit sur targomo.com/developers. Astuce : ajoute TARGOMO_API_KEY "
                 "dans les variables d'environnement Render pour ne plus la saisir.",
        )
        targomo_modes = st.multiselect(
            "Modes de transport",
            options=list(apis.TRAVEL_MODES.keys()),
            default=["walk"],
            format_func=lambda m: apis.TRAVEL_MODES[m],
        )
        targomo_times = st.multiselect(
            "Durées (minutes)",
            options=[5, 10, 15, 20, 30],
            default=[5, 10, 15],
        )
        run_targomo = st.button("Calculer les isochrones",
                                 disabled=not (target and targomo_key))

    with st.expander("🏛️ DVF (transactions notariales)", expanded=False):
        dvf_radius_km = st.slider("Rayon DVF (km)", 0.5, 5.0, 2.0, 0.5)
        dvf_years = st.slider("Années à inclure", 1, 10, 5)
        dvf_types = st.multiselect(
            "Types de locaux",
            options=list(apis.LOCAL_TERTIAIRE),
            default=list(apis.LOCAL_TERTIAIRE),
        )
        run_dvf = st.button("Charger les transactions DVF",
                             disabled=not target)

if mode == "Charger un fichier Excel":
    if not file:
        st.info("Charge ton fichier Excel dans la barre latérale pour démarrer.")
        st.stop()
    try:
        df = pd.read_excel(file, sheet_name="COMPARABLES", skiprows=2)
    except Exception:
        df = pd.read_excel(file, skiprows=2)
    n_before = len(df)
    df = df[df["Commune"].notna() & df["Surface_m2"].notna()].reset_index(drop=True)
    has_price = df["Loyer_facial_eur_m2_an"].apply(to_num).notna() | df["Prix_vente_eur_m2"].apply(to_num).notna()
    df = df[has_price].reset_index(drop=True)
    if len(df) < n_before:
        st.caption(f"ℹ️ {n_before - len(df)} ligne(s) écartée(s) (commune, surface ou prix manquant).")
    st.write(f"**{len(df)}** référence(s) exploitable(s) chargée(s) depuis le fichier.")
else:
    if "live_results" not in st.session_state:
        st.session_state.live_results = []
    if not target:
        st.info("Renseigne d'abord l'adresse ou la commune de l'actif (section 1) pour activer "
                 "la recherche en direct.")
        st.stop()
    search_commune = target["city"] or manual_commune.strip() or "Nantes"
    search_cp = target["postcode"] or manual_cp.strip() or "44000"
    if run_live:
        ops_to_search = ["Location", "Vente"] if op.startswith("Tous") else [op]
        types_to_search = ["Bureaux", "Activités", "Commerce"] if asset_type == "Tous" else [asset_type]
        n_sources = len(selected_sources) if selected_sources else len(live_search.ALL_SOURCES)
        est_sec = n_sources * max_pages * 0.5
        with st.spinner(f"Recherche en direct — {max_pages} page(s) × {n_sources} source(s) "
                         f"(~{est_sec:.0f}s)…"):
            collected = []
            for o in ops_to_search:
                for t in types_to_search:
                    collected += live_search.search_all_sources(
                        search_commune, search_cp, o, t,
                        sources=selected_sources,
                        max_pages=max_pages,
                    )

        # ── diagnostic précis ────────────────────────────────────────────
        merged = live_search.dedupe_listings(collected)
        n_brut        = len(merged)
        n_avec_prix   = sum(1 for x in merged
                            if to_num(x.get("Loyer_facial_eur_m2_an"))
                            or to_num(x.get("Prix_vente_eur_m2")))
        n_avec_addr   = sum(1 for x in merged if x.get("Adresse_precise"))
        n_geocodable  = sum(1 for x in merged
                            if x.get("Commune") not in ("Non disponible", "", None))
        n_complets    = sum(1 for x in merged if x.get("Complete"))
        filtered      = live_search.filter_complete(merged) if only_complete else merged

        # stockage
        st.session_state.live_results = filtered
        st.session_state.live_diag = {
            "brut": n_brut, "prix": n_avec_prix,
            "addr": n_avec_addr, "geocodable": n_geocodable,
            "complets": n_complets, "filtres": len(filtered),
            "only_complete": only_complete,
        }

        if not filtered:
            diag = st.session_state.live_diag
            st.error("❌ Aucun comparable retenu — voici pourquoi :")
            st.markdown(f"""
| Étape | Résultat |
|---|---|
| Annonces brutes scrappées | **{diag['brut']}** |
| Dont avec un prix extrait | **{diag['prix']}** |
| Dont avec adresse de rue précise | **{diag['addr']}** |
| Dont géocodables (commune identifiée) | **{diag['geocodable']}** |
| Dont complètes (commune + prix) | **{diag['complets']}** |
| Après filtre « complètes seulement » | **{diag['filtres']}** |
""")
            if diag["brut"] == 0:
                st.warning("**Cause probable : JavaScript.** Le site charge ses annonces via JS — "
                            "requests ne peut pas les lire. Décoche toutes les sources sauf "
                            "BureauxLocaux et réessaie.")
            elif diag["prix"] == 0:
                st.warning("**Cause probable : format de prix non reconnu.** "
                            "Décoche « annonces complètes » pour voir les résultats partiels.")
            elif diag["complets"] == 0 and only_complete:
                st.warning(f"**Filtre trop strict.** {diag['geocodable']} annonces sont "
                            f"géocodables mais sans prix associé, ou inversement. "
                            f"**Décoche « Ne garder que les annonces complètes ».**")
            st.stop()
        else:
            n_raw = len(merged)
            msg = f"✅ {len(filtered)} annonce(s) retenue(s) sur {search_commune}"
            if only_complete and n_raw > len(filtered):
                msg += f" ({n_raw - len(filtered)} écartée(s) — adresse ou prix manquant)"
            st.success(msg)
    if not st.session_state.live_results:
        st.info("Lance la recherche en direct dans la barre latérale.")
        st.stop()

    # ── Enrichissement photos ────────────────────────────────────────────
    n_sans_photo = sum(1 for r in st.session_state.live_results if not r.get("Photo_url"))
    if n_sans_photo > 0:
        n_to_fetch = min(n_sans_photo, 40)
        prog_p = st.progress(0.0, text=f"Récupération des photos ({n_to_fetch} fiches)…")
        st.session_state.live_results = live_search.fetch_photos_batch(
            st.session_state.live_results,
            max_items=n_to_fetch,
            progress_cb=lambda i, n: prog_p.progress(i / max(n, 1)),
        )
        prog_p.empty()
        n_found = sum(1 for r in st.session_state.live_results if r.get("Photo_url"))
        st.caption(f"📷 {n_found}/{len(st.session_state.live_results)} photo(s) récupérée(s).")

    # ── Enrichissement géo : PLU + Cadastre pour chaque comparable ───────
    n_sans_plu = sum(1 for r in st.session_state.live_results if not r.get("PLU_zone"))
    if n_sans_plu > 0:
        n_enrich = min(n_sans_plu, 30)
        prog_g = st.progress(0.0, text=f"Enrichissement PLU + Cadastre ({n_enrich} adresses)…")
        st.session_state.live_results = apis.enrich_comparables_batch(
            st.session_state.live_results,
            max_items=n_enrich,
            progress_cb=lambda i, n: prog_g.progress(i / max(n, 1)),
        )
        prog_g.empty()
        n_plu = sum(1 for r in st.session_state.live_results if r.get("PLU_zone"))
        if n_plu > 0:
            st.caption(f"🏙️ {n_plu}/{len(st.session_state.live_results)} comparable(s) avec zone PLU.")

        # ── Autosave automatique : le dossier se sauvegarde sans clic requis ──
        if st.session_state.get("active_dossier_id"):
            _ok = dossiers.save_dossier_data(
                st.session_state.active_dossier_id,
                dossiers.snapshot_session_state(st.session_state),
                adresse=address or st.session_state.get("address_value", ""),
            )
            if _ok:
                st.caption("💾 Dossier sauvegardé automatiquement.")

    df = pd.DataFrame(st.session_state.live_results)
    st.write(f"**{len(df)}** référence(s) retenue(s) (en direct, à valider).")
    st.dataframe(df, use_container_width=True, hide_index=True)

if not target:
    st.warning("Renseigne et valide une adresse cible pour lancer l'analyse.")
    st.stop()

t_lat, t_lon, t_label = target["lat"], target["lon"], target["label"]
radius_m = radius_km * 1000

# ---------------------------------------------------------------- géocodage des comps
progress = st.progress(0.0, text="Géocodage des comparables…")
lats, lons, dists = [], [], []
for i, row_ in df.iterrows():
    lat = to_num(row_.get("Latitude"))
    lon = to_num(row_.get("Longitude"))
    if lat is None or lon is None:
        addr_val = str(row_.get("Adresse", "") or "")
        # n'inclure l'adresse que si elle contient vraiment une info de rue
        addr_clean = "" if any(x in addr_val.lower() for x in
                               ("non disponible", "non extraite", "adresse précise")) \
                     else addr_val.strip()
        commune = str(row_.get("Commune", "") or "").strip()
        cp      = str(row_.get("Code_postal", "") or "").strip()
        # construire la query en garantissant qu'on a au moins la commune
        parts = [p for p in [addr_clean, cp, commune] if p and p != "Non disponible"]
        q = " ".join(parts)
        g = geocode(q) if q else None
        if g:
            lat, lon = g["lat"], g["lon"]
        time.sleep(0.05)
    # stocker None explicitement (pas NaN) pour les coords manquantes
    lats.append(float(lat) if lat is not None else None)
    lons.append(float(lon) if lon is not None else None)
    dists.append(haversine_m(t_lat, t_lon, lat, lon)
                 if (lat is not None and lon is not None) else None)
    progress.progress((i + 1) / max(len(df), 1), text=f"Géocodage… {i+1}/{len(df)}")
progress.empty()

df["_lat"], df["_lon"], df["_dist"] = lats, lons, dists

# ---------------------------------------------------------------- filtrage + score
def _is_unknown(v):
    return v is None or (isinstance(v, float) and pd.isna(v)) or str(v).strip() in ("", "Non disponible")

op_filter = None if op.startswith("Tous") else op
type_filter = None if asset_type == "Tous" else asset_type

def passes(row_):
    if row_["_dist"] is None or row_["_dist"] > radius_m:
        return False
    # une ligne dont l'opération/le type est inconnu n'est PAS exclue : on ne sait pas, donc on
    # ne rejette pas — seul un désaccord confirmé exclut la ligne.
    if op_filter and not _is_unknown(row_.get("Operation")) and row_.get("Operation") != op_filter:
        return False
    if type_filter and not _is_unknown(row_.get("Type_actif")) and row_.get("Type_actif") != type_filter:
        return False
    return True

res = df[df.apply(passes, axis=1)].copy()

# ── Diagnostic non-bloquant (sans st.stop) ───────────────────────────────
if res.empty:
    n_total = len(df)
    n_geocoded = int(df["_lat"].notna().sum())
    n_in_radius = int((df["_dist"].notna() & (df["_dist"] <= radius_m)).sum())
    st.warning(
        f"⚠️ 0 comparable dans le rayon {radius_km} km (sur {n_geocoded} géocodés). "
        f"Les tabs restent accessibles — élargis le rayon ou consulte la carte pour voir "
        f"où sont les annonces."
    )

# ── Score ────────────────────────────────────────────────────────────────
now = datetime.now()
def score_row(row_):
    p_prox = max(0.0, 1 - row_["_dist"] / radius_m) if row_["_dist"] else 0.5
    sc = to_num(row_.get("Surface_m2"))
    p_sim = (min(sc, surface_target) / max(sc, surface_target)
             if surface_target and sc else 0.5)
    p_rec = 0.5
    dt = pd.to_datetime(row_.get("Date_transaction"), errors="coerce")
    if pd.notna(dt):
        months = (now - dt.to_pydatetime()).days / 30.4
        p_rec = max(0.0, 1 - months / 36)
    wsum = (w_prox + w_sim + w_rec) or 1
    return (w_prox * p_prox + w_sim * p_sim + w_rec * p_rec) / wsum

if not res.empty:
    res["_score"] = res.apply(score_row, axis=1)
    res = res.sort_values("_score", ascending=False)

# ── Champ de valorisation ────────────────────────────────────────────────
if op == "Vente":
    field = "Prix_vente_eur_m2"
elif op == "Location":
    field = "Loyer_facial_eur_m2_an"
else:
    if not res.empty:
        n_vente = res["Prix_vente_eur_m2"].apply(to_num).notna().sum() if "Prix_vente_eur_m2" in res else 0
        n_location = res["Loyer_facial_eur_m2_an"].apply(to_num).notna().sum() if "Loyer_facial_eur_m2_an" in res else 0
        field = "Prix_vente_eur_m2" if n_vente >= n_location else "Loyer_facial_eur_m2_an"
    else:
        field = "Loyer_facial_eur_m2_an"

is_vente_field = field == "Prix_vente_eur_m2"
bounds = ((live_search.PRIX_M2_MIN, live_search.PRIX_M2_MAX) if is_vente_field
          else (live_search.LOYER_M2_AN_MIN, live_search.LOYER_M2_AN_MAX))

def remove_outliers_iqr(series: pd.Series, k: float = 1.5) -> pd.Series:
    """Supprime les outliers par la méthode IQR (Tukey).
    k=1.5 → outliers modérés ; k=3.0 → outliers extrêmes seulement.
    Robuste aux distributions asymétriques (loyers immo)."""
    if len(series) < 4:
        return series  # trop peu de points pour détecter des outliers
    q1, q3 = series.quantile(0.25), series.quantile(0.75)
    iqr = q3 - q1
    if iqr == 0:
        return series  # distribution plate, pas d'outliers
    lower = q1 - k * iqr
    upper = q3 + k * iqr
    return series[(series >= lower) & (series <= upper)]


if not res.empty:
    raw_vals = res[field].apply(to_num)
    plausible_mask = raw_vals.notna() & raw_vals.between(*bounds)
    vals_plaus = raw_vals[plausible_mask]
    # Filtrage IQR : retire les outliers statistiques
    vals = remove_outliers_iqr(vals_plaus)
    n_implausible = int((raw_vals.notna() & ~plausible_mask).sum())
    n_outliers = len(vals_plaus) - len(vals)
    weights = res.loc[vals.index, "_score"]
else:
    vals = pd.Series(dtype=float)
    weights = pd.Series(dtype=float)
    n_implausible = 0

unit = "€/m²" if is_vente_field else "€/m²/an HT-HC"

if n_implausible:
    st.caption(f"⚠️ {n_implausible} valeur(s) hors fourchette écartée(s).")
if not res.empty and "n_outliers" in dir():
    if n_outliers > 0:
        st.caption(f"📊 {n_outliers} outlier(s) retiré(s) par filtrage IQR (valeurs aberrantes).")

COLS_KEEP = {
    "Adresse":                "Adresse",
    "Commune":                "Commune",
    "Type_actif":             "Type",
    "Etat":                   "État",
    "Surface_m2":             "Surface (m²)",
    "Loyer_facial_eur_m2_an": "Loyer facial €/m²/an",
    "Charges_eur_m2_an":      "Charges €/m²/an",
    "Loyer_HT_HC_eur_m2_an":  "Loyer HT-HC €/m²/an",
    "Prix_vente_eur_m2":      "Prix vente €/m²",
    "Nb_parkings":            "Parkings",
    "Date_transaction":       "Date",
    "Source":                 "Source",
    "Lien":                   "Lien",
    "Photo_url":              "Photo",
    "Commentaire":            "Commentaire",
}
COL_CONFIG = {
    "Lien":                   st.column_config.LinkColumn("Lien", display_text="🔗 Annonce"),
    "Photo":                  st.column_config.LinkColumn("Photo", display_text="🖼️ Voir"),
    "Surface (m²)":           st.column_config.NumberColumn(format="%d m²"),
    "Loyer facial €/m²/an":  st.column_config.NumberColumn(format="%.0f €"),
    "Charges €/m²/an":       st.column_config.NumberColumn(format="%.0f €"),
    "Loyer HT-HC €/m²/an":   st.column_config.NumberColumn(format="%.0f €"),
    "Prix vente €/m²":        st.column_config.NumberColumn(format="%.0f €"),
    "Parkings":               st.column_config.NumberColumn(format="%d"),
    "Commentaire":            st.column_config.TextColumn(
        "Commentaire",
        help="Ex : Preneur : XYZ. Bail 3/6/9 en date du 15/01/2025.",
        width="large",
    ),
}

def valo_bloc(df_subset: pd.DataFrame, label: str, color: str, key_suffix: str):
    """Affiche un tableau éditable + les métriques de valorisation pour un sous-groupe."""
    avail = {k: v for k, v in COLS_KEEP.items() if k in df_subset.columns}
    display = df_subset[list(avail.keys())].copy().rename(columns=avail)
    if "Loyer HT-HC €/m²/an" not in display.columns:
        display["Loyer HT-HC €/m²/an"] = None

    # ── champ de valorisation : HT-HC si dispo, sinon facial en repli ──
    if is_vente_field:
        v_field = "Prix vente €/m²"
        display["_val"] = display[v_field].apply(to_num) if v_field in display.columns else None
    else:
        hthc = display["Loyer HT-HC €/m²/an"].apply(to_num) if "Loyer HT-HC €/m²/an" in display.columns else pd.Series(dtype=float)
        facial = display["Loyer facial €/m²/an"].apply(to_num) if "Loyer facial €/m²/an" in display.columns else pd.Series(dtype=float)
        # combine : HT-HC en priorité, facial quand HT-HC est absent
        combined = hthc.combine_first(facial) if not hthc.empty else facial
        display["_val"] = combined
        v_field = "_val"

    raw = display[v_field].apply(to_num) if v_field != "_val" else display["_val"]
    plaus = raw[raw.between(*bounds)]
    # Filtrage IQR : retire les outliers avant calcul des métriques
    plaus = remove_outliers_iqr(plaus)
    n = len(plaus)

    if n == 0:
        st.info(f"Aucun comparable {label} avec valeur exploitable dans le rayon.")
        return display

    w = res.loc[df_subset.index, "_score"].reindex(plaus.index).fillna(0.5)
    cible = (plaus * w).sum() / w.sum() if w.sum() > 0 else plaus.mean()
    basse, haute, med = plaus.min(), plaus.max(), plaus.median()
    conf = "🟢" if n >= 8 else ("🟡" if n >= 4 else "🔴")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Loyer bas" if not is_vente_field else "Prix bas",
              f"{basse:,.0f} {unit}".replace(",", " "))
    c2.metric("Loyer cible" if not is_vente_field else "Prix cible",
              f"{cible:,.0f} {unit}".replace(",", " "))
    c3.metric("Loyer haut" if not is_vente_field else "Prix haut",
              f"{haute:,.0f} {unit}".replace(",", " "))
    c4.metric("Comparables", n)
    st.caption(f"Médiane : {med:,.0f} {unit} — Fiabilité : {conf} ({n} comp.)".replace(",", " "))
    if n < 4:
        st.warning("Peu de comparables — élargis le rayon ou ajoute des références.")

    # tableau éditable
    st.caption("Modifiable directement — les corrections sont répercutées dans l'export.")
    edited = st.data_editor(
        display, use_container_width=True, hide_index=True,
        num_rows="fixed", column_config=COL_CONFIG, key=f"editor_{key_suffix}"
    )

    # recalcul si modification
    if v_field in edited.columns:
        new_v = edited[v_field].apply(to_num).dropna()
        new_v = new_v[new_v.between(*bounds)]
        if len(new_v) > 0:
            cible_edit = new_v.mean()
            if abs(cible_edit - cible) > 0.5:
                st.info(f"💡 Après corrections → loyer cible **{cible_edit:,.0f} {unit}**".replace(",", " "))
    return edited


# ── Split Neuf/Restructuré  vs  Seconde main ──────────────────────────────
ETATS_NEUF = {"Neuf", "Restructuré"}
ETATS_SM   = {"Seconde main"}

def etat_groupe(row):
    e = str(row.get("Etat", "") or "").strip()
    if e in ETATS_NEUF: return "neuf_restructure"
    if e in ETATS_SM:   return "seconde_main"
    return "inconnu"

if not res.empty:
    res["_groupe"] = res.apply(etat_groupe, axis=1)
    res_neuf = res[res["_groupe"] == "neuf_restructure"].copy()
    res_sm   = res[res["_groupe"] == "seconde_main"].copy()
    res_inc  = res[res["_groupe"] == "inconnu"].copy()
else:
    res["_groupe"] = pd.Series(dtype=str)
    res_neuf = res_sm = res_inc = pd.DataFrame(columns=res.columns)

total_neuf, total_sm, total_inc = len(res_neuf), len(res_sm), len(res_inc)

if not res.empty:
    st.caption(f"**{len(res)} comparables dans le rayon** : "
               f"{total_neuf} Neuf/Restructuré · {total_sm} Seconde main · "
               f"{total_inc} état non déterminé")
else:
    # Carte de la zone scrappée (hors rayon) pour aider l'utilisateur
    st.info(
        f"💡 **0 comparable dans le rayon de {radius_km} km.** "
        f"La carte ci-dessous montre les {len(df)} annonces trouvées — "
        f"élargis le rayon dans la barre latérale pour les inclure."
    )

# ── Onglets principaux ────────────────────────────────────────────────────

def _valid_coord(v) -> bool:
    """Retourne True si v est un float fini (exclut None, NaN, Inf)."""
    try:
        return v is not None and math.isfinite(float(v))
    except (TypeError, ValueError):
        return False


def make_map(target_lat: float, target_lon: float, target_label: str,
             res_df: pd.DataFrame, field: str, bounds: tuple) -> folium.Map:
    """Carte Folium centrée sur l'actif cible avec tous les comparables."""

    # centrage : on ne garde que les coordonnées finies pour éviter les NaN Folium
    valid_lats = [float(v) for v in res_df["_lat"] if _valid_coord(v)]
    valid_lons = [float(v) for v in res_df["_lon"] if _valid_coord(v)]
    all_lats = [target_lat] + valid_lats
    all_lons = [target_lon] + valid_lons
    center = [sum(all_lats) / len(all_lats), sum(all_lons) / len(all_lons)]

    m = folium.Map(location=center, zoom_start=14, tiles=None)

    # ── couches de fond ──────────────────────────────────────────────────
    folium.TileLayer(
        tiles="https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
        attr="Esri World Imagery",
        name="🛰️ Satellite",
        overlay=False, control=True,
    ).add_to(m)

    folium.TileLayer(
        tiles="https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png",
        attr="CartoDB",
        name="🗺️ Plan",
        overlay=False, control=True,
    ).add_to(m)

    # ── Overlay PLU (zones d'urbanisme) — IGN data.geopf.fr WMS ─────────
    # Couche optionnelle : zones U/AU/N/A colorées selon le PLU en vigueur
    try:
        folium.WmsTileLayer(
            url="https://data.geopf.fr/wms-r/wms?",
            name="🏙️ Zones PLU",
            fmt="image/png",
            layers="URBANISMES_REGLEMENTES.ZONE_URBA",
            transparent=True,
            version="1.3.0",
            attr="Géoportail de l'Urbanisme / IGN",
            overlay=True,
            control=True,
            opacity=0.35,
            show=False,  # désactivé par défaut, activable dans la légende
        ).add_to(m)

        # Overlay bâti (emprise des bâtiments) pour situer les immeubles
        folium.WmsTileLayer(
            url="https://data.geopf.fr/wms-r/wms?",
            name="🏛️ Bâti IGN",
            fmt="image/png",
            layers="BDTOPO_V3:batiment",
            transparent=True,
            version="1.3.0",
            attr="BD TOPO / IGN",
            overlay=True,
            control=True,
            opacity=0.5,
            show=False,
        ).add_to(m)
    except Exception:
        pass  # WMS non disponible → carte fonctionne sans overlay

    folium.LayerControl(position="topright", collapsed=True).add_to(m)

    # ── marqueur actif cible ─────────────────────────────────────────────
    folium.Marker(
        location=[target_lat, target_lon],
        popup=folium.Popup(f"<b>🏢 ACTIF À VALORISER</b><br>{target_label}", max_width=250),
        tooltip=f"🏢 {target_label}",
        icon=folium.Icon(color="red", icon="home", prefix="fa"),
    ).add_to(m)

    # rayon de recherche
    radius_m_val = radius_km * 1000
    folium.Circle(
        location=[target_lat, target_lon],
        radius=radius_m_val,
        color="#B23A48", fill=True, fill_opacity=0.04,
        tooltip=f"Rayon {radius_km:.1f} km",
    ).add_to(m)

    # ── marqueurs comparables ─────────────────────────────────────────────
    # couleur par état
    COULEUR_ETAT = {
        "Neuf": "green", "Restructuré": "blue",
        "Seconde main": "orange", "Non disponible": "gray",
    }

    # valeur min/max pour dégradé
    vals = res_df[field].apply(to_num).dropna()
    v_min, v_max = (vals.min(), vals.max()) if len(vals) >= 2 else (0, 1)

    for _, row in res_df.iterrows():
        lat, lon = row.get("_lat"), row.get("_lon")
        if not _valid_coord(lat) or not _valid_coord(lon):
            continue
        lat, lon = float(lat), float(lon)

        val = to_num(row.get(field))
        etat = str(row.get("Etat") or "Non disponible").strip()
        adresse = str(row.get("Adresse") or "").strip()
        commune = str(row.get("Commune") or "").strip()
        surface = row.get("Surface_m2")
        lien = str(row.get("Lien") or "")
        dist = row.get("_dist")
        score = row.get("_score", 0)
        source = str(row.get("Source") or "")

        # label loyer/prix
        if val is not None:
            unit_short = "€/m²" if field == "Prix_vente_eur_m2" else "€/m²/an"
            val_str = f"{val:,.0f} {unit_short}".replace(",", "\u202f")
        else:
            val_str = "Prix non extrait"

        # ── PLU, Cadastre, Risques depuis les données enrichies ──────────
        plu_zone  = str(row.get("PLU_zone", "") or "").strip()
        plu_desc  = str(row.get("PLU_desc", "") or row.get("PLU_libelle", "") or "").strip()
        cad_ref   = str(row.get("Cadastre_ref", "") or "").strip()
        cad_cont  = row.get("Cadastre_contenance")   # surface parcelle m²

        plu_html = ""
        if plu_zone:
            plu_color = apis.PLU_ZONE_COLORS.get(plu_zone[:1], "#666")
            plu_html = (f'<span style="color:{plu_color};font-weight:bold">'
                        f'PLU : {plu_zone}</span>'
                        + (f' — {plu_desc}' if plu_desc and plu_desc != plu_zone else "")
                        + "<br>")
        cad_html = ""
        if cad_ref:
            cad_html = (f'<span style="color:#888;font-size:11px">📐 {cad_ref}'
                        + (f' · parcelle {int(cad_cont)} m²' if cad_cont else "")
                        + '</span><br>')

        # popup HTML enrichi
        lien_html = (f'<a href="{lien}" target="_blank" style="color:#008493">🔗 Annonce</a>'
                     if lien and lien.startswith("http") else "")
        popup_html = f"""
        <div style="font-family:Arial,sans-serif;font-size:13px;min-width:220px;max-width:300px">
          <b style="color:#25273A">{adresse}{', ' + commune if commune else ''}</b><br>
          <span style="color:#666">{etat} · {int(surface) if surface else '?'} m²</span><br>
          <span style="font-size:16px;font-weight:bold;color:#008493">{val_str}</span><br>
          {plu_html}{cad_html}
          <span style="color:#aaa;font-size:11px">
            {f'{int(dist)} m' if dist else ''} · {source}
          </span><br>{lien_html}
        </div>
        """
        # tooltip : inclut la zone PLU si disponible
        plu_tooltip = f" · PLU {plu_zone}" if plu_zone else ""
        tooltip_str = f"{val_str} · {adresse or commune}{plu_tooltip}"
        couleur = COULEUR_ETAT.get(etat, "gray")

        folium.CircleMarker(
            location=[lat, lon],
            radius=8 if val is not None else 6,
            color="white", weight=1.5,
            fill=True, fill_color=couleur, fill_opacity=0.85,
            popup=folium.Popup(popup_html, max_width=280),
            tooltip=tooltip_str,
        ).add_to(m)

    # ── légende ─────────────────────────────────────────────────────────
    legend_html = """
    <div style="position:fixed;bottom:24px;right:24px;z-index:999;background:white;
         padding:10px 14px;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,.2);
         font-family:Arial,sans-serif;font-size:12px">
      <b>État des locaux</b><br>
      <span style="color:green">●</span> Neuf<br>
      <span style="color:blue">●</span> Restructuré<br>
      <span style="color:orange">●</span> Seconde main<br>
      <span style="color:gray">●</span> Non déterminé<br>
      <span style="color:red">⌂</span> Actif cible
    </div>
    """
    m.get_root().html.add_child(folium.Element(legend_html))

    return m


tab_analyse, tab_carte, tab_env, tab_dvf, tab_targomo, tab_urba = st.tabs([
    "📊 Analyse & comparables",
    "🗺️ Carte",
    "🏙️ Environnement",
    "🏛️ DVF — Transactions",
    "🎯 Targomo — Isochrones",
    "📋 Urbanisme",
])

st.markdown("""
<style>
[data-testid="stTabsContent"] > div { padding-top: 0 !important; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════
# ONGLET 1 — Analyse
# ══════════════════════════════════════════════════════════════════════════
with tab_analyse:
    all_edited = {}

    st.subheader("🟢 Neuf / Restructuré")
    if res_neuf.empty:
        st.info("Aucun comparable Neuf ou Restructuré dans le rayon avec ces critères.")
    else:
        all_edited["neuf"] = valo_bloc(res_neuf, "Neuf/Restructuré", "#2E7D5B", "neuf")

    st.subheader("🟠 Seconde main")
    if res_sm.empty:
        st.info("Aucun comparable Seconde main dans le rayon avec ces critères.")
    else:
        all_edited["sm"] = valo_bloc(res_sm, "Seconde main", "#B26A3C", "sm")

    if not res_inc.empty:
        with st.expander(f"État non déterminé ({total_inc} comp.) — à qualifier manuellement"):
            st.caption("Qualifie l'état dans la colonne État pour que ces comparables "
                        "remontent dans le bon tableau lors de la prochaine recherche.")
            avail_inc = {k: v for k, v in COLS_KEEP.items() if k in res_inc.columns}
            disp_inc = res_inc[list(avail_inc.keys())].copy().rename(columns=avail_inc)
            all_edited["inc"] = st.data_editor(
                disp_inc, use_container_width=True, hide_index=True,
                num_rows="fixed", column_config=COL_CONFIG, key="editor_inc"
            )

    st.divider()
    st.subheader("Export pour rapport")
    frames = []
    for key in ("neuf", "sm", "inc"):
        if key in all_edited:
            df_e = all_edited[key].copy()
            df_e["_groupe_export"] = {"neuf": "Neuf / Restructuré",
                                       "sm": "Seconde main",
                                       "inc": "État non déterminé"}[key]
            frames.append(df_e)
    all_display = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    if not all_display.empty:
        # ── Fix Bug SM : re-trier par l'État ÉDITÉ par l'utilisateur
        # Si quelqu'un a changé "État non déterminé" → "Seconde main" dans le tableau,
        # cet item doit apparaître dans la section Seconde main de l'export.
        def _groupe_export(row):
            e = str(row.get("État", "") or "").strip()
            if e in ("Neuf", "Restructuré"):  return "Neuf / Restructuré"
            if e == "Seconde main":           return "Seconde main"
            return row.get("_groupe_export", "État non déterminé")
        all_display["_groupe_export"] = all_display.apply(_groupe_export, axis=1)

        # s'assurer que la colonne Commentaire existe (vide par défaut si non renseignée)
        if "Commentaire" not in all_display.columns:
            all_display["Commentaire"] = ""

        edited_for_export = all_display.rename(columns={v: k for k, v in COLS_KEEP.items()})
        export_df = build_export_table(edited_for_export, operation=op)

        # aperçu sans la colonne Photo (URL brute peu lisible dans l'aperçu)
        preview = export_df.drop(columns=["Photo"], errors="ignore")
        st.dataframe(preview, use_container_width=True, hide_index=True)

        col_a, col_b = st.columns(2)
        with col_a:
            with st.spinner("Génération Excel (téléchargement des photos…)"):
                xlsx_bytes = export_to_excel_bytes(export_df, title=f"Extraction — {t_label}", operation=op)
            st.download_button("📥 Extraction (Excel)",
                xlsx_bytes, file_name="extraction_comparables.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with col_b:
            st.download_button("Détail complet (CSV)",
                all_display.to_csv(index=False).encode("utf-8"),
                file_name="comparables_detail.csv", mime="text/csv")


# ══════════════════════════════════════════════════════════════════════════
# ONGLET 2 — Carte pleine page avec comparables auto
# ══════════════════════════════════════════════════════════════════════════
with tab_carte:
    field_for_map = ("Prix_vente_eur_m2" if op == "Vente"
                     else "Loyer_HT_HC_eur_m2_an" if "Loyer_HT_HC_eur_m2_an" in df.columns
                     else "Loyer_facial_eur_m2_an")
    bounds_for_map = ((live_search.PRIX_M2_MIN, live_search.PRIX_M2_MAX) if op == "Vente"
                      else (live_search.LOYER_M2_AN_MIN, live_search.LOYER_M2_AN_MAX))

    # Si res est vide, affiche TOUTES les annonces géocodées (hors rayon)
    # pour aider l'utilisateur à comprendre pourquoi rien n'est retenu
    map_df = res if not res.empty else df[df["_lat"].apply(_valid_coord) & df["_lon"].apply(_valid_coord)].copy()
    if res.empty and not map_df.empty:
        st.caption(
            f"⚠️ Aucun comparable dans le rayon actuel ({radius_km} km). "
            f"La carte montre les **{len(map_df)} annonces scrappées** hors rayon "
            f"— élargis le curseur dans la barre latérale pour les inclure."
        )

    folium_map = make_map(t_lat, t_lon, t_label, map_df, field_for_map, bounds_for_map)

    # injection pour que la carte prenne tout l'espace vertical disponible
    map_html = folium_map._repr_html_()
    full_page = f"""
    <div style="width:100%;height:88vh;margin:0;padding:0;">
      {map_html}
    </div>
    <script>
      setTimeout(function(){{
        var iframes = parent.document.querySelectorAll('iframe');
        iframes.forEach(function(f){{
          if(f.contentDocument && f.contentDocument.querySelector('.folium-map')){{
            f.style.height='88vh';
          }}
        }});
      }}, 400);
    </script>
    """
    components.html(full_page, height=880, scrolling=False)
    st.caption(
        "🔴 Actif · 🟢 Neuf · 🔵 Restructuré · 🟠 Seconde main · ⚫ Inconnu "
        "| Clic → détails | Bouton ↗ : satellite ↔ plan"
    )


# ══════════════════════════════════════════════════════════════════════════
# ONGLET 3 — Environnement (OpenStreetMap / Overpass)
# ══════════════════════════════════════════════════════════════════════════
with tab_env:

    OVERPASS_URL = "https://overpass-api.de/api/interpreter"

    @st.cache_data(ttl=3600, show_spinner=False)
    def overpass_fetch(lat: float, lon: float, radius: int, filters: str) -> list:
        query = f"""[out:json][timeout:20];
(node{filters}(around:{radius},{lat},{lon});
 way{filters}(around:{radius},{lat},{lon}););
out center;"""
        try:
            r = requests.post(OVERPASS_URL, data={"data": query}, timeout=25)
            r.raise_for_status()
            return r.json().get("elements", [])
        except Exception:
            return []

    # dist_m supprimée (doublon) — on réutilise haversine_m définie plus haut dans le fichier
    dist_m = haversine_m

    def nearest(els, lat, lon):
        best, best_d = None, float("inf")
        for e in els:
            elat = e.get("lat") or (e.get("center") or {}).get("lat")
            elon = e.get("lon") or (e.get("center") or {}).get("lon")
            if elat and elon:
                d = dist_m(lat, lon, elat, elon)
                if d < best_d: best_d, best = d, e
        return best, best_d

    def wmin(m): return round(m / 80)

    st.subheader(f"🏙️ Environnement — {t_label}")
    st.caption("Données OpenStreetMap via Overpass API · Mise en cache 1h · Aucune clé API")

    with st.spinner("Analyse de l'environnement en cours…"):
        tram_els    = overpass_fetch(t_lat, t_lon, 1000, '["railway"="tram_stop"]')
        bus_els     = overpass_fetch(t_lat, t_lon,  800, '["highway"="bus_stop"]')
        train_els   = overpass_fetch(t_lat, t_lon, 3000, '["railway"="station"]')
        bike_els    = overpass_fetch(t_lat, t_lon,  500, '["amenity"="bicycle_rental"]')
        resto_300   = overpass_fetch(t_lat, t_lon,  300, '["amenity"~"restaurant|cafe|fast_food"]')
        resto_500   = overpass_fetch(t_lat, t_lon,  500, '["amenity"~"restaurant|cafe|fast_food"]')
        hotel_500   = overpass_fetch(t_lat, t_lon,  500, '["tourism"="hotel"]')
        bank_300    = overpass_fetch(t_lat, t_lon,  300, '["amenity"~"bank|atm"]')
        sport_1k    = overpass_fetch(t_lat, t_lon, 1000, '["leisure"~"fitness_centre|sports_centre"]')
        creche_1k   = overpass_fetch(t_lat, t_lon, 1000, '["amenity"~"childcare|kindergarten"]')
        supermarche = overpass_fetch(t_lat, t_lon,  500, '["shop"~"supermarket|convenience"]')
        pharmacy    = overpass_fetch(t_lat, t_lon,  500, '["amenity"="pharmacy"]')
        offices_500 = overpass_fetch(t_lat, t_lon,  500, '["office"]')

    # Transports
    st.markdown("### 🚌 Transports")
    nt, dt = nearest(tram_els,  t_lat, t_lon)
    nb, db = nearest(bus_els,   t_lat, t_lon)
    ng, dg = nearest(train_els, t_lat, t_lon)
    nv, dv = nearest(bike_els,  t_lat, t_lon)

    tc1, tc2, tc3, tc4 = st.columns(4)
    tc1.metric("🚋 Tram", f"{int(dt)} m" if nt else "—", f"{wmin(dt)} min" if nt else None)
    tc2.metric("🚌 Bus",  f"{int(db)} m" if nb else "—", f"{wmin(db)} min" if nb else None)
    tc3.metric("🚂 Gare", f"{dg/1000:.1f} km" if ng else "—", f"{wmin(dg)} min" if ng else None)
    tc4.metric("🚲 Vélos", f"{int(dv)} m" if nv else "—", f"{wmin(dv)} min" if nv else None)

    # Services
    st.markdown("### 🍽️ Services à proximité")
    sc1, sc2, sc3, sc4 = st.columns(4)
    sc1.metric("Restos/cafés 300 m", len(resto_300))
    sc2.metric("Restos/cafés 500 m", len(resto_500))
    sc3.metric("Hôtels 500 m", len(hotel_500))
    sc4.metric("Banques/ATM 300 m", len(bank_300))
    sc5, sc6, sc7, sc8 = st.columns(4)
    sc5.metric("Salles de sport 1 km", len(sport_1k))
    sc6.metric("Crèches 1 km", len(creche_1k))
    sc7.metric("Supermarchés 500 m", len(supermarche))
    sc8.metric("Pharmacies 500 m", len(pharmacy))

    # Environnement tertiaire
    st.markdown("### 🏢 Environnement tertiaire")
    st.metric("Immeubles de bureaux / locaux 500 m", len(offices_500))

    # Score
    st.markdown("### 🏅 Score synthétique")
    sc_t = min(
        (40 if nt and dt<=300 else 25 if nt and dt<=600 else 0) +
        (20 if nb and db<=200 else 10 if nb and db<=400 else 0) +
        (25 if ng and dg<=800 else 15 if ng and dg<=1500 else 0) +
        (15 if nv and dv<=300 else 0), 100)
    sc_s = min(len(resto_300)*5 + len(hotel_500)*10 + len(bank_300)*5 +
               len(sport_1k)*8 + len(creche_1k)*5, 100)
    sc_b = min(len(offices_500)*5 + len(resto_500)*2, 100)
    sc_g = round(sc_t*0.4 + sc_s*0.35 + sc_b*0.25)
    note = ("⭐⭐⭐⭐⭐ Excellent" if sc_g>=80 else "⭐⭐⭐⭐ Très bon" if sc_g>=65
            else "⭐⭐⭐ Bon" if sc_g>=50 else "⭐⭐ Moyen" if sc_g>=35 else "⭐ Faible")
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Transports /100", sc_t)
    c2.metric("Services /100", sc_s)
    c3.metric("Tertiaire /100", sc_b)
    c4.metric("Score global /100", sc_g, note)

    # Carte environnement
    st.markdown("### 🗺️ Carte des commodités")
    em = folium.Map(location=[t_lat, t_lon], zoom_start=15, tiles=None)
    folium.TileLayer(
        "https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
        attr="Esri", name="🛰️ Satellite", overlay=False, control=True).add_to(em)
    folium.TileLayer(
        "https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png",
        attr="CartoDB", name="🗺️ Plan", overlay=False, control=True).add_to(em)
    folium.LayerControl(position="topright", collapsed=False).add_to(em)
    folium.Marker([t_lat, t_lon], tooltip="🏢 Actif",
                  icon=folium.Icon(color="red", icon="home", prefix="fa")).add_to(em)

    LAYERS = [
        ("blue",      tram_els[:15],    "🚋 Tram"),
        ("cadetblue", bus_els[:30],     "🚌 Bus"),
        ("orange",    resto_500[:40],   "🍽️ Resto"),
        ("purple",    hotel_500[:10],   "🏨 Hôtel"),
        ("green",     sport_1k[:10],    "🏋️ Sport"),
        ("darkgreen", bank_300[:10],    "🏦 Banque"),
    ]
    for color, els, label in LAYERS:
        for e in els:
            elat = e.get("lat") or (e.get("center") or {}).get("lat")
            elon = e.get("lon") or (e.get("center") or {}).get("lon")
            name = e.get("tags", {}).get("name", label)
            if elat and elon:
                folium.CircleMarker([elat, elon], radius=7,
                    color="white", weight=1,
                    fill=True, fill_color=color, fill_opacity=0.85,
                    tooltip=name).add_to(em)

    components.html(em._repr_html_(), height=520, scrolling=False)
    st.caption("🔴 Actif · 🔵 Tram · 🟦 Bus · 🟠 Restos · 🟣 Hôtels · 🟢 Sport · 💚 Banques")


# ══════════════════════════════════════════════════════════════════════════
# ONGLET 4 — DVF : transactions notariales réelles
# ══════════════════════════════════════════════════════════════════════════
with tab_dvf:
    st.subheader("🏛️ Transactions DVF — Données notariales publiques")
    st.caption(
        "Source : Demandes de Valeurs Foncières — data.gouv.fr (Etalab). "
        "Transactions réelles signées chez le notaire. Ventes uniquement, pas de locations. "
        "Données gratuites, sans clé API, avec un délai de publication de ~6 mois."
    )

    if not target:
        st.info("Renseigne l'adresse de l'actif (section 1) pour charger les transactions DVF.")
    else:
        search_commune_dvf = target.get("city", "") or manual_commune.strip() or "Nantes"

        if "dvf_results" not in st.session_state:
            st.session_state.dvf_results = pd.DataFrame()

        if run_dvf:
            with st.spinner(f"Téléchargement DVF pour {search_commune_dvf}…"):
                try:
                    dvf_df = apis.get_dvf_transactions(
                        lat=t_lat, lon=t_lon,
                        radius_m=dvf_radius_km * 1000,
                        commune=search_commune_dvf,
                        types_locaux=set(dvf_types) if dvf_types else None,
                        annees=dvf_years,
                    )
                    st.session_state.dvf_results = dvf_df
                    if dvf_df.empty:
                        st.warning(
                            f"Aucune transaction trouvée dans un rayon de {dvf_radius_km} km "
                            f"pour {search_commune_dvf}. Essaie d'élargir le rayon ou de "
                            f"changer la commune dans la barre latérale."
                        )
                    else:
                        st.success(f"✅ {len(dvf_df)} transaction(s) trouvée(s).")
                except Exception as e:
                    st.error(f"Erreur DVF : {e}")

        dvf_df = st.session_state.dvf_results

        if not dvf_df.empty:
            # Métriques de synthèse
            stats = apis.dvf_summary(dvf_df)
            if stats:
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Transactions", stats["n"])
                c2.metric("Prix médian /m²", f"{stats['median']:,.0f} €".replace(",", "\u202f"))
                c3.metric("Prix min /m²",    f"{stats['min']:,.0f} €".replace(",", "\u202f"))
                c4.metric("Prix max /m²",    f"{stats['max']:,.0f} €".replace(",", "\u202f"))

            st.caption(
                f"Q1 : {stats.get('q1',0):,.0f} € · Q3 : {stats.get('q3',0):,.0f} €".replace(",", "\u202f")
                + f" · Moyenne : {stats.get('mean',0):,.0f} €/m²".replace(",", "\u202f")
            )

            # Tableau
            st.subheader("Détail des transactions")
            num_cols = ["Surface (m²)", "Prix total (€)", "Prix/m²", "Distance (m)"]
            col_cfg = {c: st.column_config.NumberColumn(format="%.0f") for c in num_cols if c in dvf_df.columns}
            st.dataframe(dvf_df, use_container_width=True, hide_index=True,
                         column_config=col_cfg)

            # Carte DVF
            st.subheader("Carte des transactions")
            dvf_map = folium.Map(location=[t_lat, t_lon], zoom_start=14, tiles=None)
            folium.TileLayer(
                "https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
                attr="Esri", name="🛰️ Satellite", overlay=False, control=True
            ).add_to(dvf_map)
            folium.TileLayer(
                "https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png",
                attr="CartoDB", name="🗺️ Plan", overlay=False, control=True
            ).add_to(dvf_map)
            folium.LayerControl(position="topright", collapsed=False).add_to(dvf_map)

            folium.Marker([t_lat, t_lon], tooltip="🏢 Actif",
                          icon=folium.Icon(color="red", icon="home", prefix="fa")).add_to(dvf_map)

            for _, row in dvf_df.iterrows():
                if not _valid_coord(row.get("_lat")) or not _valid_coord(row.get("_lon")):
                    continue
                px = row.get("Prix/m²")
                popup_html = (
                    f"<b>{row.get('Adresse','—')}, {row.get('Commune','')}</b><br>"
                    f"Date : {row.get('Date','—')}<br>"
                    f"Type : {row.get('Type','—')}<br>"
                    f"Surface : {row.get('Surface (m²)','—')} m²<br>"
                    f"<b>Prix : {str(round(px)) + ' €/m²' if px else '—'}</b>"
                )
                folium.CircleMarker(
                    location=[float(row["_lat"]), float(row["_lon"])],
                    radius=9, color="white", weight=1.5,
                    fill=True, fill_color="#008493", fill_opacity=0.85,
                    popup=folium.Popup(popup_html, max_width=240),
                    tooltip=f"DVF — {str(round(px)) + ' €/m²' if px else 'prix inconnu'}",
                ).add_to(dvf_map)

            components.html(dvf_map._repr_html_(), height=480, scrolling=False)

            # Export
            st.download_button(
                "📥 Exporter les transactions DVF (CSV)",
                dvf_df.drop(columns=["_lat","_lon"], errors="ignore")
                      .to_csv(index=False).encode("utf-8"),
                file_name="dvf_transactions.csv",
                mime="text/csv",
            )


# ══════════════════════════════════════════════════════════════════════════
# ONGLET 5 — Targomo : isochrones de déplacement
# ══════════════════════════════════════════════════════════════════════════
with tab_targomo:
    st.subheader("🎯 Isochrones Targomo — Zones accessibles par mode de transport")
    st.caption(
        "Affiche les zones atteignables depuis l'actif en N minutes selon le mode de transport. "
        "Nécessite une clé API Targomo gratuite — inscription sur "
        "[targomo.com/developers](https://targomo.com/developers/)."
    )

    if not target:
        st.info("Renseigne l'adresse de l'actif (section 1) pour calculer les isochrones.")
    elif not targomo_key:
        st.warning(
            "Renseigne ta clé API Targomo dans la section 4 de la barre latérale. "
            "La clé gratuite ('Free tier') permet jusqu'à 1 000 requêtes/mois."
        )
    else:
        if "targomo_results" not in st.session_state:
            st.session_state.targomo_results = {}

        if run_targomo:
            for mode in (targomo_modes or ["walk"]):
                with st.spinner(f"Calcul isochrones — {apis.TRAVEL_MODES.get(mode, mode)}…"):
                    result = apis.fetch_isochrones(
                        lat=t_lat, lon=t_lon,
                        api_key=targomo_key,
                        times_minutes=sorted(targomo_times or [5, 10, 15]),
                        mode=mode,
                    )
                    if result and "error" in result:
                        st.error(f"Erreur Targomo ({mode}): {result['error']}")
                    else:
                        st.session_state.targomo_results[mode] = result
                        st.success(f"✅ Isochrones calculées — {apis.TRAVEL_MODES.get(mode, mode)}")
                time.sleep(0.3)

        results = st.session_state.targomo_results
        if results:
            # Construction de la carte avec isochrones
            iso_map = folium.Map(location=[t_lat, t_lon], zoom_start=13, tiles=None)
            folium.TileLayer(
                "https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
                attr="Esri", name="🛰️ Satellite", overlay=False, control=True
            ).add_to(iso_map)
            folium.TileLayer(
                "https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png",
                attr="CartoDB", name="🗺️ Plan", overlay=False, control=True
            ).add_to(iso_map)
            folium.LayerControl(position="topright", collapsed=False).add_to(iso_map)

            for mode, geojson_data in results.items():
                if geojson_data and "error" not in geojson_data:
                    apis.add_isochrones_to_map(
                        iso_map, geojson_data,
                        times_minutes=sorted(targomo_times or [5, 10, 15]),
                        mode=mode,
                    )

            folium.Marker(
                [t_lat, t_lon], tooltip="🏢 Actif cible",
                icon=folium.Icon(color="red", icon="home", prefix="fa"),
            ).add_to(iso_map)

            components.html(iso_map._repr_html_(), height=680, scrolling=False)

            # Tableau de lisibilité des temps
            st.subheader("📏 Lecture des isochrones")
            for mode in results:
                st.markdown(f"**{apis.TRAVEL_MODES.get(mode, mode)}**")
                for t in sorted(targomo_times or [5, 10, 15]):
                    color = apis.TARGOMO_COLORS.get(
                        min(apis.TARGOMO_COLORS.keys(), key=lambda x: abs(x - t)), "#888")
                    st.markdown(
                        f"<span style='color:{color};font-size:18px'>●</span> "
                        f"Zone accessible en **{t} min** à {apis.TRAVEL_MODES.get(mode,mode)}",
                        unsafe_allow_html=True
                    )
        else:
            st.info("Clique sur 'Calculer les isochrones' dans la section 4 de la barre latérale.")

# ══════════════════════════════════════════════════════════════════════════
# ONGLET 6 — Urbanisme : PLU, cadastre, risques de l'actif cible + comparables
# ══════════════════════════════════════════════════════════════════════════
with tab_urba:
    st.subheader("📋 Urbanisme & réglementaire")
    st.caption(
        "Zone PLU (Géoportail de l'Urbanisme), référence cadastrale (IGN) et "
        "risques (Géorisques) — données publiques gratuites, sans clé API."
    )

    if not target:
        st.info("Renseigne l'adresse de l'actif (section 1) pour charger les données d'urbanisme.")
    else:
        # ── Actif cible ──────────────────────────────────────────────────
        st.markdown("### 🏢 Actif cible")
        with st.spinner("Interrogation PLU / Cadastre / Géorisques…"):
            plu_target = apis.fetch_plu_zone(t_lat, t_lon)
            cad_target = apis.fetch_cadastre_info(t_lat, t_lon)
            risques_target = apis.fetch_georisques(t_lat, t_lon)

        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("**🏙️ Zone PLU**")
            if plu_target:
                zone = plu_target.get("zone_type", "—")
                color = apis.PLU_ZONE_COLORS.get(zone[:1] if zone else "", "#666")
                st.markdown(
                    f"<span style='font-size:24px;font-weight:700;color:{color}'>{zone or '—'}</span>",
                    unsafe_allow_html=True)
                st.caption(plu_target.get("zone_description", "") or plu_target.get("zone_libelle", ""))
                if plu_target.get("libelong"):
                    st.caption(f"_{plu_target['libelong']}_")
            else:
                st.caption("Zone PLU non trouvée pour ce point. "
                           "Le PLU de la commune n'est peut-être pas encore numérisé sur le GPU.")
        with c2:
            st.markdown("**📐 Cadastre**")
            if cad_target:
                st.markdown(f"**{cad_target.get('ref_cadastrale', '—')}**")
                st.caption(f"Section {cad_target.get('section','—')} · "
                           f"Parcelle {cad_target.get('numero','—')}")
                if cad_target.get("contenance"):
                    st.caption(f"Contenance : {int(cad_target['contenance']):,} m²".replace(",", "\u202f"))
            else:
                st.caption("Parcelle non identifiée à ces coordonnées.")
        with c3:
            st.markdown("**⚠️ Risques**")
            if risques_target:
                for _, info in risques_target.items():
                    st.caption(f"• {info.get('label', '')}")
            else:
                st.caption("Aucun indicateur de risque retourné (ou API indisponible).")

        # ── Comparables : synthèse PLU ────────────────────────────────────
        st.divider()
        st.markdown("### 📊 Zones PLU des comparables")
        if res.empty or "PLU_zone" not in res.columns:
            st.info("Lance d'abord une recherche de comparables — les zones PLU sont "
                     "récupérées automatiquement pour chaque adresse géocodée.")
        else:
            plu_series = res["PLU_zone"].fillna("").replace("", "Non déterminé")
            plu_counts = plu_series.value_counts()

            cc1, cc2 = st.columns([1, 2])
            with cc1:
                st.markdown("**Répartition**")
                for zone, count in plu_counts.items():
                    color = apis.PLU_ZONE_COLORS.get(str(zone)[:1], "#666")
                    st.markdown(
                        f"<span style='color:{color};font-weight:700'>{zone}</span> : {count} comp.",
                        unsafe_allow_html=True)
            with cc2:
                st.markdown("**Détail par comparable**")
                cols_urba = [c for c in ["Adresse", "Commune", "PLU_zone", "PLU_desc",
                                          "Cadastre_ref", "Cadastre_contenance"] if c in res.columns]
                if cols_urba:
                    disp_urba = res[cols_urba].copy().rename(columns={
                        "PLU_zone": "Zone PLU", "PLU_desc": "Description",
                        "Cadastre_ref": "Réf. cadastrale",
                        "Cadastre_contenance": "Parcelle (m²)",
                    })
                    st.dataframe(disp_urba, use_container_width=True, hide_index=True)

            # Note méthodologique
            st.caption(
                "💡 La zone PLU influence directement la valeur : un comparable en zone UE "
                "(économique) n'est pas strictement comparable à un actif en zone UC (centre-ville). "
                "Vérifie la cohérence des zones avant de valider la fourchette de valorisation."
            )
